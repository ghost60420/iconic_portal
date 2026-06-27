import io
import json
import logging
from collections import defaultdict
from decimal import Decimal

from django.conf import settings
from django.contrib import messages
from django.core.files.base import ContentFile
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST

from .forms_costing import CostingHeaderForm, CostingSMVForm, OpportunityDocumentForm, QuickCostingForm
from .models import (
    CostingHeader,
    CostingLineItem,
    CostingSMV,
    CostingAuditLog,
    CostingSnapshot,
    NEW_COSTING_CATEGORY_CHOICES,
    NEW_COSTING_CURRENCY_CHOICES,
    NEW_COSTING_UOM_CHOICES,
    Opportunity,
    OpportunityDocument,
    QuickCosting,
)
from .services.costing_currency import format_costing_money, normalize_costing_currency
from .services.costing_engine import compute_costing, validate_costing
from .services.costing_workflow import (
    CostingWorkflowError,
    convert_costing_to_quotation,
    create_invoice_from_costing,
    create_invoice_from_quick_costing,
    get_costing_quote_amounts,
)
from .services.order_lifecycle import create_lifecycle_from_costing
from .services.workflow_visibility import build_workflow_visibility_context
from .permissions import can_view_internal_costing


logger = logging.getLogger(__name__)


DEFAULT_QUOTATION_TERMS = """For bulk orders, 50% advance confirms the order and 50% is due before shipment.

For samples, 100% payment is required before development begins.

Production starts after payment is cleared.

Any change after approval may affect price and timeline.

Shipping time may vary due to courier, customs, or international delay.

Import duties and local taxes are the buyer's responsibility unless agreed otherwise.

Any issue must be reported within 5 days of receiving goods.

All agreements are governed under the laws of British Columbia, Canada."""


def _can_approve(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    access = getattr(user, "access", None)
    return bool(access and access.can_costing_approve)


def _can_convert_to_invoice(user):
    return bool(user and user.is_authenticated and user.is_superuser)


def _user_or_none(user):
    return user if user and getattr(user, "is_authenticated", False) else None


def _next_quick_quotation_number():
    prefix = f"QQT{timezone.now():%Y}"
    latest = (
        QuickCosting.objects.filter(quotation_number__startswith=prefix)
        .exclude(quotation_number="")
        .order_by("-quotation_number")
        .first()
    )
    next_num = 1
    if latest and latest.quotation_number:
        try:
            next_num = int(latest.quotation_number.replace(prefix, "")) + 1
        except ValueError:
            next_num = 1

    for offset in range(1000):
        candidate = f"{prefix}{next_num + offset:04}"
        if not QuickCosting.objects.filter(quotation_number=candidate).exists():
            return candidate

    return f"{prefix}{timezone.now():%m%d%H%M%S}"


def _costing_currency(costing):
    return normalize_costing_currency(getattr(costing, "currency", None))


def _format_costing_money(costing, value):
    return format_costing_money(value, _costing_currency(costing))


def _format_quick_decimal(value):
    value = value or Decimal("0")
    return f"{value.quantize(Decimal('0.01')):,.2f}"


def _format_quick_percent(value):
    value = value or Decimal("0")
    return f"{value.quantize(Decimal('0.01')):,.2f}"


def _format_quick_bdt(value):
    return f"৳{_format_quick_decimal(value)} BDT"


def _format_quick_currency(value, currency):
    currency = (currency or "BDT").upper()
    if currency == "CAD":
        return f"CAD ${_format_quick_decimal(value)}"
    if currency == "USD":
        return f"USD ${_format_quick_decimal(value)}"
    return _format_quick_bdt(value)


def _format_quick_cad_from_bdt(value, exchange_rate):
    if not exchange_rate:
        return "CAD N/A"
    try:
        cad_value = (value or Decimal("0")) / exchange_rate
    except Exception:
        return "CAD N/A"
    return f"CAD ${_format_quick_decimal(cad_value)}"


def _format_quick_money_pair(value, exchange_rate, currency="BDT", is_legacy_currency=True):
    if not is_legacy_currency:
        return _format_quick_currency(value, currency)
    return f"{_format_quick_bdt(value)} / {_format_quick_cad_from_bdt(value, exchange_rate)}"


def _format_quick_money_lines(value, exchange_rate, currency="BDT", is_legacy_currency=True):
    if not is_legacy_currency:
        return {
            "bdt": _format_quick_currency(value, currency),
            "cad": "",
        }
    return {
        "bdt": _format_quick_bdt(value),
        "cad": _format_quick_cad_from_bdt(value, exchange_rate),
    }


def _quick_costing_calc(quick_costing):
    summary = quick_costing.calculation_summary()
    exchange_rate = summary.get("exchange_rate")
    currency = summary["currency"]
    is_legacy_currency = summary["is_legacy_currency"]
    calc = {
        "total_cost_order": summary["total_cost"],
        "total_sales_order": summary["revenue"],
        "total_profit_order": summary["total_profit"],
        "total_cost_per_piece": summary["cost_per_piece"],
        "fob_per_piece": summary["selling_price_per_piece"],
        "profit_per_piece": summary["profit_per_piece"],
        "margin_percent": summary["profit_margin_percent"],
        "quantity": summary["quantity"],
        "currency": currency,
        "is_legacy_currency": is_legacy_currency,
        "exchange_rate": exchange_rate,
        "uses_detailed_costing": summary["uses_detailed_costing"],
        "fabric_cost_per_kg": summary["fabric_cost_per_kg"],
        "fabric_consumption_kg_per_piece": summary["fabric_consumption_kg_per_piece"],
        "fabric_cost_per_piece": summary["fabric_cost_per_piece"],
        "making_cost_per_piece": summary["making_cost_per_piece"],
        "print_embroidery_cost_per_piece": summary["print_embroidery_cost_per_piece"],
        "trims_cost_per_piece": summary["trims_cost_per_piece"],
        "packaging_cost_per_piece": summary["packaging_cost_per_piece"],
        "material_cost": summary["material_cost_total"],
        "material_cost_per_piece": summary["material_cost_per_piece"],
        "material_cost_total": summary["material_cost_total"],
        "production_cost": summary["production_cost_total"],
        "production_cost_per_piece": summary["production_cost_per_piece"],
        "production_cost_total": summary["production_cost_total"],
        "other_expenses": summary["other_expenses_total"],
        "other_expenses_per_piece": summary["other_expenses_per_piece"],
        "other_expenses_total": summary["other_expenses_total"],
        "shipping_cost": summary["shipping_cost_total"],
        "shipping_cost_per_piece": summary["shipping_cost_per_piece"],
        "shipping_cost_total": summary["shipping_cost_total"],
        "selling_price_per_piece": summary["selling_price_per_piece"],
        "selling_price_total": summary["selling_price_total"],
        "gross_profit_per_piece": summary["gross_profit_per_piece"],
        "gross_profit_total": summary["gross_profit_total"],
        "commission_per_piece": summary["commission_per_piece"],
        "commission_total": summary["commission_total"],
        "commission_percent": summary["commission_percent"],
        "net_profit_per_piece": summary["net_profit_per_piece"],
        "net_profit_total": summary["net_profit_total"],
        "gross_profit_margin_percent": summary["gross_profit_margin_percent"],
        "net_profit_margin_percent": summary["net_profit_margin_percent"],
        "target_margin_percent": summary["target_margin_percent"],
        "margin_status": summary["margin_status"],
    }
    money_pair = lambda value: _format_quick_money_pair(value, exchange_rate, currency, is_legacy_currency)
    money_lines = lambda value: _format_quick_money_lines(value, exchange_rate, currency, is_legacy_currency)
    calc["display"] = {
        "total_cost_order": _format_quick_decimal(calc["total_cost_order"]),
        "total_cost_order_pair": money_pair(calc["total_cost_order"]),
        "total_cost_order_lines": money_lines(calc["total_cost_order"]),
        "total_sales_order": _format_quick_decimal(calc["total_sales_order"]),
        "total_sales_order_pair": money_pair(calc["total_sales_order"]),
        "total_sales_order_lines": money_lines(calc["total_sales_order"]),
        "total_profit_order": _format_quick_decimal(calc["total_profit_order"]),
        "total_profit_order_pair": money_pair(calc["total_profit_order"]),
        "total_profit_order_lines": money_lines(calc["total_profit_order"]),
        "total_cost_per_piece": _format_quick_decimal(calc["total_cost_per_piece"]),
        "total_cost_per_piece_pair": money_pair(calc["total_cost_per_piece"]),
        "total_cost_per_piece_lines": money_lines(calc["total_cost_per_piece"]),
        "fob_per_piece": _format_quick_decimal(calc["fob_per_piece"]),
        "fob_per_piece_pair": money_pair(calc["fob_per_piece"]),
        "fob_per_piece_lines": money_lines(calc["fob_per_piece"]),
        "profit_per_piece": _format_quick_decimal(calc["profit_per_piece"]),
        "profit_per_piece_pair": money_pair(calc["profit_per_piece"]),
        "margin_percent": _format_quick_decimal(calc["margin_percent"]),
        "fabric_cost_per_kg": money_pair(calc["fabric_cost_per_kg"]),
        "fabric_consumption_kg_per_piece": f"{calc['fabric_consumption_kg_per_piece'].quantize(Decimal('0.0001'))} kg / piece",
        "fabric_cost_per_piece": money_pair(calc["fabric_cost_per_piece"]),
        "making_cost_per_piece": money_pair(calc["making_cost_per_piece"]),
        "print_embroidery_cost_per_piece": money_pair(calc["print_embroidery_cost_per_piece"]),
        "trims_cost_per_piece": money_pair(calc["trims_cost_per_piece"]),
        "packaging_cost_per_piece": money_pair(calc["packaging_cost_per_piece"]),
        "material_cost": _format_quick_decimal(calc["material_cost"]),
        "material_cost_per_piece_pair": money_pair(calc["material_cost_per_piece"]),
        "material_cost_total_pair": money_pair(calc["material_cost_total"]),
        "production_cost": _format_quick_decimal(calc["production_cost"]),
        "production_cost_per_piece_pair": money_pair(calc["production_cost_per_piece"]),
        "production_cost_total_pair": money_pair(calc["production_cost_total"]),
        "other_expenses": _format_quick_decimal(calc["other_expenses"]),
        "other_expenses_per_piece_pair": money_pair(calc["other_expenses_per_piece"]),
        "other_expenses_total_pair": money_pair(calc["other_expenses_total"]),
        "shipping_cost": _format_quick_decimal(calc["shipping_cost"]),
        "shipping_cost_per_piece_pair": money_pair(calc["shipping_cost_per_piece"]),
        "shipping_cost_total_pair": money_pair(calc["shipping_cost_total"]),
        "selling_price_per_piece_pair": money_pair(calc["selling_price_per_piece"]),
        "selling_price_total_pair": money_pair(calc["selling_price_total"]),
        "gross_profit_per_piece": _format_quick_decimal(calc["gross_profit_per_piece"]),
        "gross_profit_per_piece_pair": money_pair(calc["gross_profit_per_piece"]),
        "gross_profit_total": _format_quick_decimal(calc["gross_profit_total"]),
        "gross_profit_total_pair": money_pair(calc["gross_profit_total"]),
        "commission_per_piece": _format_quick_decimal(calc["commission_per_piece"]),
        "commission_per_piece_pair": money_pair(calc["commission_per_piece"]),
        "commission_total": _format_quick_decimal(calc["commission_total"]),
        "commission_total_pair": money_pair(calc["commission_total"]),
        "commission_percent": _format_quick_percent(calc["commission_percent"]) if calc["commission_percent"] is not None else "N/A",
        "commission_percent_label": f"{_format_quick_percent(calc['commission_percent'])}%" if calc["commission_percent"] is not None else "Legacy absolute amount",
        "net_profit_per_piece": _format_quick_decimal(calc["net_profit_per_piece"]),
        "net_profit_per_piece_pair": money_pair(calc["net_profit_per_piece"]),
        "net_profit_total": _format_quick_decimal(calc["net_profit_total"]),
        "net_profit_total_pair": money_pair(calc["net_profit_total"]),
        "net_profit_total_lines": money_lines(calc["net_profit_total"]),
        "gross_profit_margin_percent": _format_quick_percent(calc["gross_profit_margin_percent"]),
        "net_profit_margin_percent": _format_quick_percent(calc["net_profit_margin_percent"]),
        "target_margin_percent": _format_quick_percent(calc["target_margin_percent"]) if calc["target_margin_percent"] is not None else "N/A",
        "target_margin_percent_label": f"{_format_quick_percent(calc['target_margin_percent'])}%" if calc["target_margin_percent"] is not None else "N/A",
        "margin_status": calc["margin_status"],
        "currency": "Legacy BDT with CAD conversion" if is_legacy_currency else currency,
        "exchange_rate": f"1 CAD = {_format_quick_decimal(exchange_rate)} BDT" if exchange_rate else "N/A",
    }
    return calc


def _deny_without_internal_costing(request):
    if can_view_internal_costing(request.user):
        return None
    return HttpResponseForbidden("No access")


def _quotation_company():
    return {
        "name": getattr(settings, "INVOICE_COMPANY_NAME", "Iconic Apparel House"),
        "email": getattr(settings, "INVOICE_COMPANY_EMAIL", "info@iconicapparelhouse.com"),
        "phone": getattr(settings, "INVOICE_COMPANY_PHONE", "604-500-6009"),
        "website": getattr(settings, "INVOICE_COMPANY_WEBSITE", "iconicapparelhouse.com"),
        "address": getattr(settings, "INVOICE_ADDRESS_CA", ""),
    }


def _display_user(user):
    if not user:
        return ""
    return user.get_full_name() or user.get_username()


def _quotation_ceo_status_label(costing, converted=False):
    if converted:
        return "Converted"
    if costing.quotation_status == CostingHeader.QUOTATION_STATUS_APPROVED:
        return "CEO Approved"
    if costing.quotation_status == CostingHeader.QUOTATION_STATUS_REJECTED:
        return "CEO Rejected"
    if costing.quotation_status == CostingHeader.QUOTATION_STATUS_SENT:
        return "Sent to Client"
    if costing.quotation_number:
        return "Submitted for CEO Approval"
    return "Draft"


def _quotation_salesperson_label(costing):
    if costing.quoted_by:
        return _display_user(costing.quoted_by)
    if costing.approved_by:
        return _display_user(costing.approved_by)
    lead = getattr(getattr(costing, "opportunity", None), "lead", None)
    if getattr(lead, "assigned_to", None):
        return _display_user(lead.assigned_to)
    if getattr(lead, "owner", ""):
        return lead.owner
    return ""


def _quick_approval_status_label(quick_costing):
    if getattr(quick_costing, "rejected_at", None) or quick_costing.status == QuickCosting.STATUS_REJECTED:
        return "Rejected"
    if getattr(quick_costing, "approved_at", None) or quick_costing.status in {
        QuickCosting.STATUS_APPROVED,
        QuickCosting.STATUS_QUOTED,
        QuickCosting.STATUS_INVOICED,
        QuickCosting.STATUS_PRODUCTION,
        QuickCosting.STATUS_SHIPPED,
        QuickCosting.STATUS_CLOSED,
    }:
        return "Approved"
    return "Pending"


def _quick_profit_health(calc):
    margin = calc.get("net_profit_margin_percent") or Decimal("0")
    if margin > Decimal("30"):
        return {"label": "Excellent", "tone": "excellent"}
    if margin >= Decimal("20"):
        return {"label": "Healthy", "tone": "healthy"}
    if margin >= Decimal("10"):
        return {"label": "Low Margin", "tone": "low"}
    return {"label": "Loss", "tone": "loss"}


def _quick_workflow_badges(quick_costing, invoice=None):
    status = quick_costing.status
    if invoice or status in {
        QuickCosting.STATUS_INVOICED,
        QuickCosting.STATUS_PRODUCTION,
        QuickCosting.STATUS_SHIPPED,
        QuickCosting.STATUS_CLOSED,
    }:
        current_key = "invoiced"
    elif status == QuickCosting.STATUS_QUOTED:
        current_key = "quoted"
    elif status == QuickCosting.STATUS_REJECTED:
        current_key = "rejected"
    elif status == QuickCosting.STATUS_APPROVED:
        current_key = "ceo-approved"
    else:
        current_key = "ceo-pending"

    stages = [
        ("draft", "Draft"),
        ("submitted", "Submitted"),
        ("ceo-pending", "CEO Pending"),
        ("ceo-approved", "CEO Approved"),
        ("rejected", "Rejected"),
        ("quoted", "Quoted"),
        ("invoiced", "Invoiced"),
    ]
    normal_order = ["draft", "submitted", "ceo-pending", "ceo-approved", "quoted", "invoiced"]
    current_index = normal_order.index(current_key) if current_key in normal_order else -1
    badges = []
    for key, label in stages:
        if key == current_key:
            state = "current"
        elif key in normal_order and current_index >= 0 and normal_order.index(key) < current_index:
            state = "complete"
        else:
            state = "inactive"
        badges.append({"key": key, "label": label, "state": state})
    return badges


def _quotation_context(costing, user=None):
    amounts = get_costing_quote_amounts(costing)
    quote_is_approved = costing.quotation_status == CostingHeader.QUOTATION_STATUS_APPROVED
    converted = costing.invoices.exists()
    return {
        "costing": costing,
        "amounts": amounts,
        "company": _quotation_company(),
        "terms": DEFAULT_QUOTATION_TERMS,
        "can_approve_quotation": _can_approve(user),
        "can_convert_to_invoice": _can_convert_to_invoice(user) and quote_is_approved,
        "user_can_convert_to_invoice": _can_convert_to_invoice(user),
        "quote_is_approved": quote_is_approved,
        "quotation_ceo_status_label": _quotation_ceo_status_label(costing, converted=converted),
    }


def _workflow_context(costing, user=None):
    invoice = costing.invoices.select_related("order").order_by("-created_at", "-id").first()
    production_order = None
    if invoice and invoice.order_id:
        production_order = invoice.order
    if not production_order:
        production_order = costing.production_orders.order_by("-created_at", "-id").first()
    lifecycle = costing.order_lifecycles_as_quotation.order_by("-updated_at", "-id").first()
    if not lifecycle:
        lifecycle = costing.order_lifecycles_as_costing.order_by("-updated_at", "-id").first()
    return {
        "is_quotation": bool(
            costing.quotation_number
            and costing.quoted_at
            and costing.quotation_status == CostingHeader.QUOTATION_STATUS_APPROVED
        ),
        "invoice": invoice,
        "production_order": production_order,
        "lifecycle": lifecycle,
        "can_convert_to_invoice": _can_convert_to_invoice(user),
    }


def _pdf_lines(pdf, text, max_width, font_name="Helvetica", font_size=9):
    words = (text or "").split()
    if not words:
        return [""]

    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if pdf.stringWidth(candidate, font_name, font_size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def _safe_costing_smv(costing):
    try:
        return costing.smv
    except CostingSMV.DoesNotExist:
        return None


def _build_line_dict(line=None, category=None):
    if line:
        return {
            "id": line.id,
            "category": line.category,
            "item_name": line.item_name,
            "item_reference": line.item_reference,
            "supplier": line.supplier,
            "uom": line.uom,
            "unit_price": line.unit_price,
            "freight": line.freight,
            "consumption_value": line.consumption_value,
            "wastage_percent": line.wastage_percent,
            "denominator_value": line.denominator_value if line.denominator_value is not None else "",
            "placement": line.placement,
            "color": line.color,
            "gsm": line.gsm,
            "cut_width": line.cut_width,
            "remarks": line.remarks,
            "sort_order": line.sort_order,
        }
    return {
        "id": None,
        "category": category or "other",
        "item_name": "",
        "item_reference": "",
        "supplier": "",
        "uom": "piece",
        "unit_price": "",
        "freight": "0",
        "consumption_value": "1",
        "wastage_percent": "0",
        "denominator_value": "1",
        "placement": "",
        "color": "",
        "gsm": "",
        "cut_width": "",
        "remarks": "",
        "sort_order": 0,
    }


def _parse_line_payload(raw):
    if raw is None:
        return []
    raw = raw.strip()
    if not raw:
        return []
    try:
        payload = json.loads(raw)
    except Exception:
        return []
    if not isinstance(payload, list):
        return []
    return payload


def _save_line_items(costing, payload):
    seen_ids = []
    for idx, row in enumerate(payload, start=1):
        item_name = (row.get("item_name") or "").strip()
        if not item_name:
            continue
        line_id = row.get("id")
        line = None
        if line_id:
            line = CostingLineItem.objects.filter(pk=line_id, costing=costing).first()
        if not line:
            line = CostingLineItem(costing=costing)
        line.category = (row.get("category") or "other").strip() or "other"
        line.item_name = item_name
        line.item_reference = (row.get("item_reference") or "").strip()
        line.supplier = (row.get("supplier") or "").strip()
        line.uom = (row.get("uom") or "piece").strip() or "piece"
        line.unit_price = row.get("unit_price") or 0
        line.freight = row.get("freight") or 0
        line.consumption_value = row.get("consumption_value") or 1
        line.wastage_percent = row.get("wastage_percent") or 0
        line.denominator_value = row.get("denominator_value") or 1
        line.placement = (row.get("placement") or "").strip()
        line.color = (row.get("color") or "").strip()
        line.gsm = (row.get("gsm") or "").strip()
        line.cut_width = (row.get("cut_width") or "").strip()
        line.remarks = (row.get("remarks") or "").strip()
        line.sort_order = int(row.get("sort_order") or idx)
        line.save()
        seen_ids.append(line.id)

    CostingLineItem.objects.filter(costing=costing).exclude(id__in=seen_ids).delete()


def _group_lines(costing):
    grouped = {key: [] for key, _ in NEW_COSTING_CATEGORY_CHOICES}
    for line in costing.line_items.all().order_by("category", "sort_order", "id"):
        grouped[line.category].append(_build_line_dict(line))
    for key in grouped:
        if not grouped[key]:
            grouped[key].append(_build_line_dict(category=key))
    return grouped


def _update_opportunity_summary(costing, calc):
    opp = costing.opportunity
    opp.costing_total_cost_per_piece = calc["total_cost_per_piece"]
    opp.costing_fob_per_piece = calc["fob_per_piece"]
    opp.costing_margin_percent = calc["margin_percent"]
    opp.costing_status = costing.status
    opp.save(update_fields=[
        "costing_total_cost_per_piece",
        "costing_fob_per_piece",
        "costing_margin_percent",
        "costing_status",
    ])


def _costing_header_initial(opportunity=None):
    initial = {
        "factory_location": "bd",
        "currency": "BDT",
        "costing_date": timezone.now().date(),
    }
    if not opportunity:
        return initial

    initial.update(
        {
            "opportunity": opportunity,
            "customer": opportunity.customer,
            "product_type": opportunity.product_type,
            "order_quantity": opportunity.moq_units or 0,
            "moq": opportunity.moq_units or 0,
            "brand": getattr(opportunity.lead, "account_brand", "") or "",
        }
    )
    return initial


def _opportunity_account_snapshot(opportunity):
    if not opportunity:
        return "", ""
    customer = getattr(opportunity, "customer", None)
    lead = getattr(opportunity, "lead", None)
    account_brand = (
        getattr(customer, "account_brand", "")
        or getattr(lead, "account_brand", "")
        or ""
    )
    contact_name = (
        getattr(customer, "contact_name", "")
        or getattr(lead, "contact_name", "")
        or ""
    )
    return account_brand, contact_name


def _quick_costing_initial(opportunity=None):
    if not opportunity:
        return {}
    account_brand, _contact_name = _opportunity_account_snapshot(opportunity)
    quantity = opportunity.moq_units or 1
    if quantity < 1:
        quantity = 1
    project_name = (
        opportunity.product_category
        or opportunity.opportunity_id
        or "Quick Costing"
    )
    return {
        "buyer_name": account_brand,
        "project_name": project_name,
        "product_type": opportunity.product_type or "Other",
        "quantity": quantity,
    }


def ceo_quotation_approval_queue(request):
    status_filter = (request.GET.get("status") or "pending").strip()
    currency = (request.GET.get("currency") or "").strip().upper()
    search = (request.GET.get("q") or "").strip()
    date_from = parse_date((request.GET.get("date_from") or "").strip())
    date_to = parse_date((request.GET.get("date_to") or "").strip())

    qs = (
        CostingHeader.objects.select_related(
            "opportunity",
            "opportunity__lead",
            "customer",
            "quoted_by",
            "approved_by",
            "quotation_approved_by",
            "quotation_rejected_by",
        )
        .prefetch_related("invoices")
        .exclude(quotation_number="")
        .order_by("-quoted_at", "-updated_at", "-id")
    )

    if currency in {"CAD", "USD", "BDT"}:
        qs = qs.filter(currency=currency)
    else:
        currency = ""

    if date_from:
        qs = qs.filter(quoted_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(quoted_at__date__lte=date_to)

    if status_filter == "pending":
        qs = qs.filter(quotation_status=CostingHeader.QUOTATION_STATUS_DRAFT)
    elif status_filter == "approved":
        qs = qs.filter(quotation_status=CostingHeader.QUOTATION_STATUS_APPROVED)
    elif status_filter == "rejected":
        qs = qs.filter(quotation_status=CostingHeader.QUOTATION_STATUS_REJECTED)
    elif status_filter == "sent":
        qs = qs.filter(quotation_status=CostingHeader.QUOTATION_STATUS_SENT)
    elif status_filter == "converted":
        qs = qs.filter(invoices__isnull=False).distinct()
    elif status_filter != "all":
        status_filter = "pending"
        qs = qs.filter(quotation_status=CostingHeader.QUOTATION_STATUS_DRAFT)

    if search:
        qs = qs.filter(
            Q(quotation_number__icontains=search)
            | Q(opportunity__opportunity_id__icontains=search)
            | Q(opportunity__lead__lead_id__icontains=search)
            | Q(customer__account_brand__icontains=search)
            | Q(customer__contact_name__icontains=search)
            | Q(style_name__icontains=search)
        )

    rows = []
    for costing in qs[:200]:
        try:
            amounts = get_costing_quote_amounts(costing)
        except CostingWorkflowError:
            amounts = {
                "order_total": Decimal("0"),
                "standard_cost_total": Decimal("0"),
            }
        total_amount = amounts["order_total"]
        profit_amount = total_amount - amounts["standard_cost_total"]
        profit_margin = (profit_amount / total_amount * Decimal("100")) if total_amount else Decimal("0")
        converted = bool(list(costing.invoices.all()))
        opportunity = getattr(costing, "opportunity", None)
        lead = getattr(opportunity, "lead", None)
        rows.append(
            {
                "costing": costing,
                "lead_id": getattr(lead, "lead_id", "") or "",
                "opportunity_id": getattr(opportunity, "opportunity_id", "") or "",
                "client_name": (
                    getattr(costing.customer, "account_brand", "")
                    or getattr(costing.customer, "contact_name", "")
                    or getattr(lead, "account_brand", "")
                    or "Client"
                ),
                "salesperson": _quotation_salesperson_label(costing) or "Not assigned",
                "currency": costing.currency or "BDT",
                "total_amount": total_amount,
                "profit_amount": profit_amount,
                "profit_margin": profit_margin,
                "purpose": "Sample" if costing.order_quantity and costing.order_quantity <= 5 else "Bulk",
                "submitted_at": costing.quoted_at or costing.created_at,
                "status_label": _quotation_ceo_status_label(costing, converted=converted),
                "converted": converted,
            }
        )

    context = {
        "rows": rows,
        "status_filter": status_filter,
        "currency": currency,
        "q": search,
        "date_from": date_from,
        "date_to": date_to,
        "currency_options": ["CAD", "USD", "BDT"],
        "status_options": [
            ("pending", "Submitted for CEO Approval"),
            ("approved", "CEO Approved"),
            ("rejected", "CEO Rejected"),
            ("sent", "Sent to Client"),
            ("converted", "Converted"),
            ("all", "All statuses"),
        ],
        "can_use_existing_approval": _can_approve(request.user) and can_view_internal_costing(request.user),
    }
    return render(request, "crm/costing/ceo_quotation_approval_queue.html", context)


def cost_sheet_list(request):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    can_view_costing_profit = can_view_internal_costing(request.user)
    qs = CostingHeader.objects.select_related("opportunity", "customer").order_by("-updated_at")
    quick_qs = QuickCosting.objects.select_related("created_by", "opportunity").order_by("-updated_at")

    customer_id = (request.GET.get("customer") or "").strip()
    product_type = (request.GET.get("product_type") or "").strip()
    status = (request.GET.get("status") or "").strip()
    costing_type = (request.GET.get("costing_type") or "all").strip()
    purpose = (request.GET.get("purpose") or "").strip()
    currency = (request.GET.get("currency") or "").strip().upper()
    date_from = parse_date((request.GET.get("date_from") or "").strip())
    date_to = parse_date((request.GET.get("date_to") or "").strip())
    if costing_type not in {"all", "advanced", "quick"}:
        costing_type = "all"
    search = (request.GET.get("q") or "").strip()
    if customer_id:
        qs = qs.filter(customer_id=customer_id)
        quick_qs = quick_qs.none()
    if product_type:
        qs = qs.filter(product_type=product_type)
        quick_qs = quick_qs.filter(product_type=product_type)
    if purpose:
        qs = qs.none()
        quick_qs = quick_qs.filter(costing_purpose=purpose)
    if status:
        qs = qs.filter(status=status)
        quick_qs = quick_qs.filter(status=status)
    if currency in {"CAD", "USD", "BDT"}:
        qs = qs.filter(currency=currency)
        if currency == "BDT":
            quick_qs = quick_qs.filter(Q(currency="BDT") | Q(currency__isnull=True))
        else:
            quick_qs = quick_qs.filter(currency=currency)
    else:
        currency = ""
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
        quick_qs = quick_qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
        quick_qs = quick_qs.filter(created_at__date__lte=date_to)
    if search:
        qs = qs.filter(
            Q(opportunity__opportunity_id__icontains=search)
            | Q(customer__account_brand__icontains=search)
            | Q(style_name__icontains=search)
            | Q(style_code__icontains=search)
        )
        quick_qs = quick_qs.filter(
            Q(buyer_name__icontains=search)
            | Q(project_name__icontains=search)
            | Q(product_type__icontains=search)
        )
    if costing_type == "advanced":
        quick_qs = quick_qs.none()
    elif costing_type == "quick":
        qs = qs.none()

    rows = []
    advanced_rows = []
    for sheet in qs:
        calc = compute_costing(sheet.id)
        if calc:
            margin_percent = calc.get("margin_percent") or Decimal("0")
            if margin_percent >= Decimal("20"):
                margin_tone = "good"
            elif margin_percent >= Decimal("5"):
                margin_tone = "watch"
            else:
                margin_tone = "risk"
            row = {
                "id": sheet.id,
                "sheet": sheet,
                "quick": None,
                "calc": calc,
                "margin_tone": margin_tone,
                "costing_type": "advanced",
                "type_label": "Advanced",
                "currency_label": sheet.currency,
                "created_at": sheet.created_at,
                "updated_at": sheet.updated_at,
            }
            rows.append(row)
            advanced_rows.append(row)

    for quick in quick_qs:
        calc = _quick_costing_calc(quick)
        margin_percent = calc.get("net_profit_margin_percent") or Decimal("0")
        if calc.get("net_profit_total", Decimal("0")) >= Decimal("0") and margin_percent >= Decimal("0"):
            margin_tone = "good"
        else:
            margin_tone = "risk"
        rows.append(
            {
                "id": quick.id,
                "sheet": None,
                "quick": quick,
                "calc": calc,
                "margin_tone": margin_tone,
                "costing_type": "quick",
                "type_label": "Quick",
                "currency_label": "BDT / CAD" if calc["is_legacy_currency"] else calc["currency"],
                "created_at": quick.created_at,
                "updated_at": quick.updated_at,
            }
        )

    rows.sort(key=lambda row: (row["updated_at"], row["id"]), reverse=True)

    total_cost_order = sum((row["calc"].get("total_cost_order") or Decimal("0")) for row in advanced_rows)
    total_sales_order = sum((row["calc"].get("total_sales_order") or Decimal("0")) for row in advanced_rows)
    total_profit_order = sum((row["calc"].get("total_profit_order") or Decimal("0")) for row in advanced_rows)
    summary_by_currency = defaultdict(lambda: {
        "total_cost_order": Decimal("0"),
        "total_sales_order": Decimal("0"),
        "total_profit_order": Decimal("0"),
    })
    for row in advanced_rows:
        currency = _costing_currency(row["sheet"])
        summary_by_currency[currency]["total_cost_order"] += row["calc"].get("total_cost_order") or Decimal("0")
        summary_by_currency[currency]["total_sales_order"] += row["calc"].get("total_sales_order") or Decimal("0")
        summary_by_currency[currency]["total_profit_order"] += row["calc"].get("total_profit_order") or Decimal("0")
    currency_summary_rows = [
        {
            "currency": currency,
            "total_cost_order": values["total_cost_order"],
            "total_sales_order": values["total_sales_order"],
            "total_profit_order": values["total_profit_order"],
        }
        for currency, values in sorted(summary_by_currency.items())
    ]
    margin_values = [row["calc"].get("margin_percent") or Decimal("0") for row in advanced_rows]
    average_margin = (sum(margin_values) / Decimal(len(margin_values))) if margin_values else Decimal("0")
    customers_by_id = {
        row["sheet"].customer_id: row["sheet"].customer
        for row in advanced_rows
        if row["sheet"].customer_id and row["sheet"].customer
    }

    context = {
        "rows": rows,
        "customers": sorted(
            customers_by_id.values(),
            key=lambda customer: (customer.account_brand or customer.contact_name or "").lower(),
        ),
        "status_choices": [
            ("draft", "Draft"),
            ("approved", "Approved"),
            ("rejected", "Rejected"),
            ("quoted", "Quoted"),
        ],
        "product_types": Opportunity.PRODUCT_TYPE_CHOICES,
        "purpose_choices": QuickCosting.PURPOSE_CHOICES,
        "currency_choices": NEW_COSTING_CURRENCY_CHOICES,
        "costing_type_choices": [
            ("all", "All"),
            ("advanced", "Advanced"),
            ("quick", "Quick"),
        ],
        "summary": {
            "count": len(rows),
            "advanced_count": len(advanced_rows),
            "quick_count": sum(1 for row in rows if row["costing_type"] == "quick"),
            "approved_count": sum(1 for row in advanced_rows if row["sheet"].status == "approved"),
            "draft_count": sum(1 for row in advanced_rows if row["sheet"].status == "draft"),
            "total_cost_order": total_cost_order,
            "total_sales_order": total_sales_order,
            "total_profit_order": total_profit_order,
            "average_margin": average_margin,
            "by_currency": currency_summary_rows,
        },
        "selected": {
            "customer": customer_id,
            "product_type": product_type,
            "status": status,
            "costing_type": costing_type,
            "purpose": purpose,
            "currency": currency,
            "date_from": date_from,
            "date_to": date_to,
            "q": search,
        },
        "can_view_internal_costing": can_view_costing_profit,
    }
    return render(request, "crm/costing/costsheet_list.html", context)


def cost_sheet_create(request, opportunity_id=None):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    opportunity = None
    if opportunity_id:
        opportunity = get_object_or_404(Opportunity, pk=opportunity_id)

    costing_type = (request.POST.get("costing_type") if request.method == "POST" else request.GET.get("costing_type")) or "advanced"
    if costing_type not in {"advanced", "quick"}:
        costing_type = "advanced"

    if costing_type == "quick":
        if request.method == "POST":
            quick_form = QuickCostingForm(request.POST)
            if quick_form.is_valid():
                quick_costing = quick_form.save(commit=False)
                if opportunity:
                    account_brand, contact_name = _opportunity_account_snapshot(opportunity)
                    quick_costing.opportunity = opportunity
                    quick_costing.account_brand = account_brand
                    quick_costing.contact_name = contact_name
                quick_costing.created_by = request.user if request.user.is_authenticated else None
                quick_costing.save()
                messages.success(request, "Quick costing saved.")
                return redirect("quick_costing_detail", pk=quick_costing.pk)
            messages.error(request, "Please fix the errors below.")
        else:
            quick_form = QuickCostingForm(initial=_quick_costing_initial(opportunity))

        context = {
            "quick_form": quick_form,
            "opportunity": opportunity,
            "mode": "create",
            "costing_type": "quick",
        }
        return render(request, "crm/costing/costsheet_form.html", context)

    if request.method == "POST":
        data = request.POST.copy()
        if opportunity:
            data["opportunity"] = opportunity.pk
            if opportunity.customer_id:
                data["customer"] = opportunity.customer_id
        form = CostingHeaderForm(data)
        if form.is_valid():
            costing = form.save(commit=False)
            if opportunity:
                costing.opportunity = opportunity
            costing.save()
            create_lifecycle_from_costing(costing, user=request.user)
            CostingAuditLog.objects.create(
                costing=costing,
                action="created",
                changed_by=request.user if request.user.is_authenticated else None,
            )
            messages.success(request, "Costing header created. Add line items next.")
            return redirect("cost_sheet_detail", pk=costing.pk)
        messages.error(request, "Please fix the errors below.")
    else:
        initial = _costing_header_initial(opportunity)
        form = CostingHeaderForm(initial=initial)

    if opportunity:
        if "opportunity" in form.fields:
            form.fields["opportunity"].disabled = True
        if "customer" in form.fields:
            form.fields["customer"].disabled = True

    context = {
        "form": form,
        "opportunity": opportunity,
        "mode": "create",
        "costing_type": "advanced",
    }
    return render(request, "crm/costing/costsheet_form.html", context)


def quick_costing_detail(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(
        QuickCosting.objects.select_related("created_by", "opportunity", "opportunity__lead", "approved_by", "rejected_by", "quoted_by"),
        pk=pk,
    )
    calc = _quick_costing_calc(quick_costing)
    invoice = quick_costing.invoices.select_related("customer", "order").order_by("-created_at", "-id").first()
    quick_costing._workflow_invoice_resolved = True
    margin_tone = "positive" if calc.get("net_profit_total", Decimal("0")) >= Decimal("0") else "negative"
    workflow_visibility = build_workflow_visibility_context(
        "costing",
        user=request.user,
        opportunity=quick_costing.opportunity,
        quick_costing=quick_costing,
        invoice=invoice,
    )
    context = {
        "quick_costing": quick_costing,
        "calc": calc,
        "invoice": invoice,
        "margin_tone": margin_tone,
        "profit_health": _quick_profit_health(calc),
        "quick_workflow_badges": _quick_workflow_badges(quick_costing, invoice=invoice),
        "can_approve": _can_approve(request.user),
        "can_convert_to_invoice": _can_convert_to_invoice(request.user) and _quick_approval_status_label(quick_costing) == "Approved",
        "approval_status_label": _quick_approval_status_label(quick_costing),
        "can_edit_quick_costing": (not quick_costing.is_locked) or _can_approve(request.user),
        **workflow_visibility,
    }
    return render(request, "crm/costing/quick_costing_detail.html", context)


def quick_costing_edit(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(
        QuickCosting.objects.select_related("opportunity", "opportunity__lead"),
        pk=pk,
    )
    if quick_costing.is_locked and not _can_approve(request.user):
        messages.error(request, "Approved quick costing is locked.")
        return redirect("quick_costing_detail", pk=pk)

    if request.method == "POST":
        form = QuickCostingForm(request.POST, instance=quick_costing)
        if form.is_valid():
            form.save()
            messages.success(request, "Quick costing updated.")
            return redirect("quick_costing_detail", pk=pk)
        messages.error(request, "Please fix the errors below.")
    else:
        form = QuickCostingForm(instance=quick_costing)

    context = {
        "quick_form": form,
        "quick_costing": quick_costing,
        "opportunity": quick_costing.opportunity,
        "mode": "edit",
        "costing_type": "quick",
    }
    return render(request, "crm/costing/costsheet_form.html", context)


@require_POST
def quick_costing_approve(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(QuickCosting, pk=pk)
    if not _can_approve(request.user):
        messages.error(request, "You do not have permission to approve.")
        return redirect("quick_costing_detail", pk=pk)
    quick_costing.status = QuickCosting.STATUS_APPROVED
    quick_costing.approved_by = _user_or_none(request.user)
    quick_costing.approved_at = timezone.now()
    quick_costing.rejected_by = None
    quick_costing.rejected_at = None
    quick_costing.save(update_fields=["status", "approved_by", "approved_at", "rejected_by", "rejected_at", "updated_at"])
    messages.success(request, "Quick costing approved and locked.")
    return redirect("quick_costing_detail", pk=pk)


@require_POST
def quick_costing_reject(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(QuickCosting, pk=pk)
    if not _can_approve(request.user):
        messages.error(request, "You do not have permission to reject.")
        return redirect("quick_costing_detail", pk=pk)
    if quick_costing.invoices.exists():
        messages.error(request, "This quick costing already has an invoice and cannot be rejected.")
        return redirect("quick_costing_detail", pk=pk)
    quick_costing.status = QuickCosting.STATUS_REJECTED
    quick_costing.rejected_by = _user_or_none(request.user)
    quick_costing.rejected_at = timezone.now()
    quick_costing.approved_by = None
    quick_costing.approved_at = None
    quick_costing.quotation_number = ""
    quick_costing.quoted_by = None
    quick_costing.quoted_at = None
    quick_costing.save(
        update_fields=[
            "status",
            "rejected_by",
            "rejected_at",
            "approved_by",
            "approved_at",
            "quotation_number",
            "quoted_by",
            "quoted_at",
            "updated_at",
        ]
    )
    messages.success(request, "Quick costing rejected.")
    return redirect("quick_costing_detail", pk=pk)


@require_POST
def quick_costing_convert_to_quotation(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(QuickCosting, pk=pk)
    if not _can_approve(request.user):
        messages.error(request, "You do not have permission to create a quotation.")
        return redirect("quick_costing_detail", pk=pk)
    if quick_costing.status != QuickCosting.STATUS_APPROVED and not quick_costing.quotation_number:
        messages.error(request, "Approve costing before creating quotation.")
        return redirect("quick_costing_detail", pk=pk)
    if not quick_costing.quotation_number:
        quick_costing.quotation_number = _next_quick_quotation_number()
        quick_costing.quoted_by = _user_or_none(request.user)
        quick_costing.quoted_at = timezone.now()
    quick_costing.status = QuickCosting.STATUS_QUOTED
    quick_costing.save(update_fields=["status", "quotation_number", "quoted_by", "quoted_at", "updated_at"])
    messages.success(request, f"Quick quotation {quick_costing.quotation_number} is ready.")
    return redirect("quick_costing_client_quotation", pk=pk)


def quick_costing_client_quotation(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(
        QuickCosting.objects.select_related(
            "created_by",
            "quoted_by",
            "approved_by",
            "rejected_by",
            "opportunity",
            "opportunity__lead",
        ),
        pk=pk,
    )
    if not quick_costing.quotation_number or not quick_costing.quoted_at:
        messages.error(request, "Approve and create a quotation before viewing it.")
        return redirect("quick_costing_detail", pk=pk)
    calc = _quick_costing_calc(quick_costing)
    invoice = quick_costing.invoices.select_related("customer", "order").order_by("-created_at", "-id").first()
    quick_costing._workflow_invoice_resolved = True
    quotation_total = (calc.get("total_sales_order") or Decimal("0")) + (calc.get("shipping_cost_total") or Decimal("0"))
    quotation_total_pair = _format_quick_money_pair(
        quotation_total,
        calc.get("exchange_rate"),
        calc.get("currency"),
        calc.get("is_legacy_currency", True),
    )
    workflow_visibility = build_workflow_visibility_context(
        "quotation",
        user=request.user,
        opportunity=quick_costing.opportunity,
        quick_costing=quick_costing,
        quotation=quick_costing,
        invoice=invoice,
    )
    context = {
        "quick_costing": quick_costing,
        "calc": calc,
        "invoice": invoice,
        "company": _quotation_company(),
        "prepared_by": _display_user(quick_costing.quoted_by or quick_costing.created_by),
        "approval_status_label": _quick_approval_status_label(quick_costing),
        "approval_user": _display_user(quick_costing.approved_by),
        "can_convert_to_invoice": _can_convert_to_invoice(request.user) and _quick_approval_status_label(quick_costing) == "Approved",
        "quotation_total_pair": quotation_total_pair,
        **workflow_visibility,
    }
    return render(request, "crm/costing/quick_quotation_client.html", context)


@require_POST
def quick_costing_convert_to_invoice(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(
        QuickCosting.objects.select_related("opportunity", "opportunity__lead"),
        pk=pk,
    )
    if not _can_convert_to_invoice(request.user):
        messages.error(request, "Only invoice managers can create invoices.")
        return redirect("quick_costing_client_quotation", pk=pk)
    if _quick_approval_status_label(quick_costing) != "Approved":
        messages.error(request, "Approve the quick costing quotation before creating an invoice.")
        return redirect("quick_costing_client_quotation", pk=pk)
    if not quick_costing.quotation_number or not quick_costing.quoted_at:
        messages.error(request, "Create a quick quotation before creating an invoice.")
        return redirect("quick_costing_detail", pk=pk)

    try:
        invoice, created = create_invoice_from_quick_costing(quick_costing, user=request.user)
    except CostingWorkflowError as exc:
        messages.error(request, str(exc))
        return redirect("quick_costing_client_quotation", pk=pk)

    if created:
        messages.success(request, f"Invoice {invoice.invoice_number} created from quick quotation.")
    else:
        messages.info(request, f"Invoice {invoice.invoice_number} already exists for this quick quotation.")
    return redirect("invoice_view", pk=invoice.pk)


def quick_costing_export_excel(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(
        QuickCosting.objects.select_related("created_by", "approved_by", "opportunity"),
        pk=pk,
    )
    try:
        from openpyxl import Workbook
    except Exception:
        messages.error(request, "Excel export is unavailable. Please install openpyxl.")
        return redirect("quick_costing_detail", pk=pk)

    try:
        calc = _quick_costing_calc(quick_costing)
        wb = Workbook()
        ws = wb.active
        ws.title = "Quick Costing"
        rows = [
            ("Buyer Name", quick_costing.buyer_name),
            ("Project Name", quick_costing.project_name),
            ("Product Type", quick_costing.get_product_type_display()),
            ("Quantity", f"{quick_costing.quantity} pieces"),
            ("Currency", calc["currency"]),
            ("Exchange Rate", calc["display"]["exchange_rate"]),
        ]
        if calc["uses_detailed_costing"]:
            rows.extend([
                ("Fabric Cost Per KG", f"{calc['display']['fabric_cost_per_kg']} / kg"),
                ("Fabric Consumption Per Piece", calc["display"]["fabric_consumption_kg_per_piece"]),
                ("Fabric Cost Per Piece", calc["display"]["fabric_cost_per_piece"]),
                ("Making Cost Per Piece", calc["display"]["making_cost_per_piece"]),
                ("Print or Embroidery Cost Per Piece", calc["display"]["print_embroidery_cost_per_piece"]),
                ("Trims Cost Per Piece", calc["display"]["trims_cost_per_piece"]),
                ("Packaging Cost Per Piece", calc["display"]["packaging_cost_per_piece"]),
            ])
        else:
            rows.extend([
                ("Legacy Material Cost - Total Order", calc["display"]["material_cost_total_pair"]),
                ("Legacy Production Cost - Total Order", calc["display"]["production_cost_total_pair"]),
            ])
        rows.extend([
            ("Other Expenses - Total Order", calc["display"]["other_expenses_total_pair"]),
            ("Shipping Cost - Total Order", calc["display"]["shipping_cost_total_pair"]),
            ("Total Cost", calc["display"]["total_cost_order_pair"]),
            ("Cost Per Piece", calc["display"]["total_cost_per_piece_pair"]),
            ("Selling Price Per Piece", calc["display"]["fob_per_piece_pair"]),
            ("Revenue", calc["display"]["total_sales_order_pair"]),
            ("Profit Before Commission", calc["display"]["gross_profit_total_pair"]),
            ("Commission Percent", calc["display"]["commission_percent_label"]),
            ("Commission Total", calc["display"]["commission_total_pair"]),
            ("Final Profit After Commission", calc["display"]["net_profit_total_pair"]),
            ("Gross Profit Margin %", calc["gross_profit_margin_percent"]),
            ("Net Profit Margin %", calc["net_profit_margin_percent"]),
            ("Target Margin %", quick_costing.target_margin_percent or ""),
            ("Margin Status", calc["margin_status"]),
            ("Status", quick_costing.get_status_display()),
            ("Created By", _display_user(quick_costing.created_by)),
            ("Created Date", quick_costing.created_at.strftime("%Y-%m-%d %H:%M")),
            ("Approved By", _display_user(quick_costing.approved_by)),
            ("Approved Date", quick_costing.approved_at.strftime("%Y-%m-%d %H:%M") if quick_costing.approved_at else ""),
        ])
        for label, value in rows:
            ws.append([label, value])

        output = io.BytesIO()
        wb.save(output)
        data = output.getvalue()
    except Exception:
        logger.exception("Failed to generate quick costing Excel", extra={"quick_costing": quick_costing.pk})
        messages.error(request, "Could not generate the Excel file. Please try again.")
        return redirect("quick_costing_detail", pk=pk)

    filename = f"quick_costing_{quick_costing.pk}.xlsx"
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write(data)
    return resp


def quick_costing_export_pdf(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    quick_costing = get_object_or_404(
        QuickCosting.objects.select_related("created_by", "opportunity"),
        pk=pk,
    )

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError:
        messages.error(request, "PDF export is unavailable. Please install ReportLab.")
        return redirect("quick_costing_detail", pk=pk)

    try:
        calc = _quick_costing_calc(quick_costing)
        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4, pageCompression=0)
        width, height = A4
        left = 36
        right = width - 36
        black = colors.HexColor("#111111")
        gold = colors.HexColor("#c89b3c")
        pale = colors.HexColor("#f6f6f4")
        border = colors.HexColor("#e5e7eb")
        muted = colors.HexColor("#6b7280")

        def text(value, fallback="-"):
            value = value if value not in (None, "") else fallback
            return str(value)

        def draw_logo(x, y):
            try:
                from django.contrib.staticfiles import finders

                logo_path = finders.find("img/image.png")
            except Exception:
                logo_path = None
            if logo_path:
                pdf.drawImage(logo_path, x, y - 34, width=42, height=42, preserveAspectRatio=True, mask="auto")
                return
            pdf.setStrokeColor(gold)
            pdf.setLineWidth(1.2)
            pdf.circle(x + 18, y - 14, 20, fill=0, stroke=1)
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawCentredString(x + 18, y - 18, "IAH")

        def draw_page_brand_header():
            pdf.setFillColor(black)
            pdf.rect(0, height - 102, width, 102, fill=1, stroke=0)
            pdf.setFillColor(gold)
            pdf.rect(0, height - 102, width, 5, fill=1, stroke=0)
            pdf.rect(right - 148, height - 66, 148, 1.4, fill=1, stroke=0)
            draw_logo(left, height - 25)
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 15)
            pdf.drawString(left + 52, height - 38, "Iconic Apparel House")
            pdf.setFont("Helvetica", 8.8)
            pdf.drawString(left + 52, height - 54, "Premium apparel sourcing, development, and production")
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawRightString(right, height - 38, f"QC-{quick_costing.pk}")
            pdf.setFont("Helvetica", 8.5)
            pdf.drawRightString(right, height - 54, timezone.localdate().strftime("%Y-%m-%d"))

        def draw_table_header(y_pos):
            pdf.setFillColor(gold)
            pdf.rect(left, y_pos - 22, right - left, 22, fill=1, stroke=0)
            pdf.setFillColor(black)
            pdf.setFont("Helvetica-Bold", 8.5)
            pdf.drawString(left + 8, y_pos - 14, "SL")
            pdf.drawString(left + 40, y_pos - 14, "Metric")
            currency_label = "BDT / CAD" if calc["is_legacy_currency"] else calc["currency"]
            pdf.drawRightString(right - 170, y_pos - 14, f"Per Piece - {currency_label}")
            pdf.drawRightString(right - 8, y_pos - 14, f"Total Order - {currency_label}")
            return y_pos - 22

        prepared_by = "Iconic Team"
        if quick_costing.created_by_id:
            prepared_by = quick_costing.created_by.get_full_name() or quick_costing.created_by.get_username()

        draw_page_brand_header()
        y = height - 132

        pdf.setFillColor(black)
        pdf.setFont("Helvetica-Bold", 24)
        pdf.drawCentredString(width / 2, y, "COSTING SHEET")
        y -= 26
        pdf.setFillColor(gold)
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawCentredString(width / 2, y, text(quick_costing.project_name, "Project"))
        y -= 18
        pdf.setFillColor(black)
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawCentredString(
            width / 2,
            y,
            f"{quick_costing.quantity or 0} PCS {quick_costing.get_product_type_display()}".upper(),
        )
        y -= 24

        box_width = (right - left - 12) / 2
        pdf.setFillColor(pale)
        pdf.roundRect(left, y - 42, box_width, 42, 6, fill=1, stroke=0)
        pdf.roundRect(left + box_width + 12, y - 42, box_width, 42, 6, fill=1, stroke=0)
        pdf.setFillColor(muted)
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(left + 10, y - 15, "BUYER NAME")
        pdf.drawString(left + box_width + 22, y - 15, "DATE")
        pdf.setFillColor(black)
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(left + 10, y - 31, text(quick_costing.buyer_name))
        pdf.drawString(left + box_width + 22, y - 31, quick_costing.created_at.strftime("%Y-%m-%d"))
        y -= 62

        if calc["uses_detailed_costing"]:
            cost_rows = [
                ("Fabric Cost", calc["display"]["fabric_cost_per_piece"], calc["display"]["material_cost_total_pair"]),
                ("Making and Finishing", calc["display"]["production_cost_per_piece_pair"], calc["display"]["production_cost_total_pair"]),
            ]
        else:
            cost_rows = [
                ("Material Cost", calc["display"]["material_cost_per_piece_pair"], calc["display"]["material_cost_total_pair"]),
                ("Production Cost", calc["display"]["production_cost_per_piece_pair"], calc["display"]["production_cost_total_pair"]),
            ]
        rows = cost_rows + [
            ("Other Expenses", calc["display"]["other_expenses_per_piece_pair"], calc["display"]["other_expenses_total_pair"]),
            ("Shipping Cost", calc["display"]["shipping_cost_per_piece_pair"], calc["display"]["shipping_cost_total_pair"]),
            ("Total Cost", calc["display"]["total_cost_per_piece_pair"], calc["display"]["total_cost_order_pair"]),
            ("Selling Price", calc["display"]["selling_price_per_piece_pair"], calc["display"]["selling_price_total_pair"]),
            ("Profit Before Commission", calc["display"]["gross_profit_per_piece_pair"], calc["display"]["gross_profit_total_pair"]),
            ("Commission", calc["display"]["commission_per_piece_pair"], calc["display"]["commission_total_pair"]),
            ("Final Profit After Commission", calc["display"]["net_profit_per_piece_pair"], calc["display"]["net_profit_total_pair"]),
        ]

        y = draw_table_header(y)
        for index, (label, per_piece, total_order) in enumerate(rows, start=1):
            row_height = 30
            if y - row_height < 125:
                pdf.showPage()
                draw_page_brand_header()
                y = draw_table_header(height - 132)
            pdf.setFillColor(colors.white if index % 2 else colors.HexColor("#fcfcfd"))
            pdf.rect(left, y - row_height, right - left, row_height, fill=1, stroke=0)
            pdf.setStrokeColor(border)
            pdf.line(left, y - row_height, right, y - row_height)
            pdf.setFillColor(black)
            pdf.setFont("Helvetica", 8.4)
            pdf.drawString(left + 8, y - 18, str(index))
            pdf.setFont("Helvetica-Bold", 8.8)
            pdf.drawString(left + 40, y - 18, label)
            pdf.setFont("Helvetica", 8.8)
            pdf.drawRightString(right - 170, y - 18, text(per_piece))
            pdf.drawRightString(right - 8, y - 18, text(total_order))
            y -= row_height

        summary_rows = [
            ("Buyer Name", text(quick_costing.buyer_name)),
            ("Project Name", text(quick_costing.project_name)),
            ("Product Type", text(quick_costing.get_product_type_display())),
            ("Date", quick_costing.created_at.strftime("%Y-%m-%d")),
            ("Quantity", text(quick_costing.quantity)),
            ("Currency", calc["display"]["currency"]),
            ("Exchange Rate", calc["display"]["exchange_rate"]),
            ("Total Order Value", calc["display"]["total_sales_order_pair"]),
            ("Total Cost", calc["display"]["total_cost_order_pair"]),
            ("Cost Per Piece", calc["display"]["total_cost_per_piece_pair"]),
            ("Selling Price Per Piece", calc["display"]["fob_per_piece_pair"]),
            ("Gross Profit Per Piece", calc["display"]["gross_profit_per_piece_pair"]),
            ("Gross Profit Total", calc["display"]["gross_profit_total_pair"]),
            ("Commission Per Piece", calc["display"]["commission_per_piece_pair"]),
            ("Commission Total", calc["display"]["commission_total_pair"]),
            ("Commission Percent", calc["display"]["commission_percent_label"]),
            ("Net Profit Per Piece", calc["display"]["net_profit_per_piece_pair"]),
            ("Net Profit Total", calc["display"]["net_profit_total_pair"]),
            ("Gross Profit Margin", f"{calc['display']['gross_profit_margin_percent']}%"),
            ("Net Profit Margin", f"{calc['display']['net_profit_margin_percent']}%"),
            ("Target Margin", calc["display"]["target_margin_percent_label"]),
            ("Margin Status", calc["display"]["margin_status"]),
            ("Prepared By", text(prepared_by)),
        ]

        y -= 20
        summary_box_height = 52 + (((len(summary_rows) + 1) // 2) * 23)
        if y < summary_box_height + 110:
            pdf.showPage()
            draw_page_brand_header()
            y = height - 132

        pdf.setFillColor(black)
        pdf.roundRect(left, y - summary_box_height, right - left, summary_box_height, 8, fill=1, stroke=0)
        pdf.setFillColor(gold)
        pdf.rect(left, y - 4, right - left, 4, fill=1, stroke=0)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(left + 12, y - 18, "Summary")
        summary_y = y - 42
        col_x = [left + 12, left + 286]
        for idx, (label, value) in enumerate(summary_rows):
            x = col_x[idx % 2]
            if idx and idx % 2 == 0:
                summary_y -= 23
            pdf.setFillColor(gold)
            pdf.setFont("Helvetica-Bold", 7.8)
            pdf.drawString(x, summary_y, label.upper())
            pdf.setFillColor(colors.white)
            pdf.setFont("Helvetica-Bold", 8.6)
            pdf.drawString(x, summary_y - 13, text(value)[:48])

        footer_y = 62
        pdf.setFillColor(black)
        pdf.rect(0, 0, width, footer_y + 30, fill=1, stroke=0)
        pdf.setFillColor(gold)
        pdf.rect(0, footer_y + 28, width, 3, fill=1, stroke=0)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawCentredString(width / 2, footer_y + 10, "Thank You!")
        pdf.setFillColor(gold)
        pdf.setFont("Helvetica", 10)
        pdf.drawCentredString(width / 2, footer_y - 5, "For Your Business")

        pdf.save()
        pdf_bytes = buffer.getvalue()
    except Exception:
        logger.exception("Failed to generate quick costing PDF", extra={"quick_costing": quick_costing.pk})
        messages.error(request, "Could not generate the PDF. Please try again.")
        return redirect("quick_costing_detail", pk=pk)

    filename = f"quick_costing_{quick_costing.pk}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(pdf_bytes)
    return response


def cost_sheet_detail(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    can_view_costing_profit = can_view_internal_costing(request.user)
    costing = get_object_or_404(
        CostingHeader.objects.select_related("opportunity", "customer").prefetch_related("line_items"),
        pk=pk,
    )
    can_approve = _can_approve(request.user)
    is_locked = costing.status == "approved"

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "save_costing":
            if is_locked:
                messages.error(request, "Costing is approved and locked.")
                return redirect("cost_sheet_detail", pk=pk)

            data = request.POST.copy()
            data["opportunity"] = costing.opportunity_id
            if costing.customer_id:
                data["customer"] = costing.customer_id

            form = CostingHeaderForm(data, instance=costing)
            smv_form = CostingSMVForm(data, instance=_safe_costing_smv(costing))
            if form.is_valid() and smv_form.is_valid():
                before = {
                    "order_quantity": costing.order_quantity,
                    "currency": costing.currency,
                    "exchange_rate": str(costing.exchange_rate or ""),
                    "finance_percent_fabric": str(costing.finance_percent_fabric),
                    "finance_percent_trims": str(costing.finance_percent_trims),
                    "commission_percent": str(costing.commission_percent),
                    "target_margin_percent": str(costing.target_margin_percent or ""),
                    "manual_fob_per_piece": str(costing.manual_fob_per_piece or ""),
                    "shipping_cost": str(costing.shipping_cost or ""),
                }
                header = form.save()
                smv = smv_form.save(commit=False)
                smv.costing = header
                smv.save()

                payload = _parse_line_payload(request.POST.get("line_payload"))
                _save_line_items(header, payload)

                calc = compute_costing(header.id)
                if calc:
                    _update_opportunity_summary(header, calc)
                create_lifecycle_from_costing(header, user=request.user)

                CostingAuditLog.objects.create(
                    costing=header,
                    action="updated",
                    changed_by=request.user if request.user.is_authenticated else None,
                    before_data=before,
                    after_data={
                        "order_quantity": header.order_quantity,
                        "currency": header.currency,
                        "exchange_rate": str(header.exchange_rate or ""),
                        "finance_percent_fabric": str(header.finance_percent_fabric),
                        "finance_percent_trims": str(header.finance_percent_trims),
                        "commission_percent": str(header.commission_percent),
                        "target_margin_percent": str(header.target_margin_percent or ""),
                        "manual_fob_per_piece": str(header.manual_fob_per_piece or ""),
                        "shipping_cost": str(header.shipping_cost or ""),
                    },
                )

                messages.success(request, "Costing updated.")
            else:
                messages.error(request, "Please fix the errors below.")
            return redirect("cost_sheet_detail", pk=pk)

        if action == "approve":
            if not can_approve:
                messages.error(request, "You do not have permission to approve.")
                return redirect("cost_sheet_detail", pk=pk)

            calc = compute_costing(costing.id)
            errors, warnings = validate_costing(costing, calc)
            if errors:
                for err in errors:
                    messages.error(request, err)
                return redirect("cost_sheet_detail", pk=pk)
            for warn in warnings:
                messages.warning(request, warn)

            costing.status = "approved"
            costing.approved_by = request.user if request.user.is_authenticated else None
            costing.approved_at = timezone.now()
            costing.save(update_fields=["status", "approved_by", "approved_at"])
            create_lifecycle_from_costing(costing, user=request.user)

            if calc:
                _update_opportunity_summary(costing, calc)
                CostingSnapshot.objects.create(
                    costing=costing,
                    data={
                        "total_cost_per_piece": str(calc["total_cost_per_piece"]),
                        "fob_per_piece": str(calc["fob_per_piece"]),
                        "margin_percent": str(calc["margin_percent"]),
                        "total_cost_order": str(calc["total_cost_order"]),
                        "total_sales_order": str(calc["total_sales_order"]),
                        "total_profit_order": str(calc["total_profit_order"]),
                        "breakdown": {k: str(v) for k, v in calc["breakdown"].items()},
                    },
                )

            CostingAuditLog.objects.create(
                costing=costing,
                action="approved",
                changed_by=request.user if request.user.is_authenticated else None,
            )
            messages.success(request, "Costing approved and locked.")
            return redirect("cost_sheet_detail", pk=pk)

        if action == "unlock":
            if not can_approve:
                messages.error(request, "You do not have permission to unlock.")
                return redirect("cost_sheet_detail", pk=pk)
            reason = (request.POST.get("unlock_reason") or "").strip()
            if not reason:
                messages.error(request, "Unlock reason is required.")
                return redirect("cost_sheet_detail", pk=pk)
            costing.status = "draft"
            costing.approved_by = None
            costing.approved_at = None
            costing.save(update_fields=["status", "approved_by", "approved_at"])
            CostingAuditLog.objects.create(
                costing=costing,
                action="unlocked",
                changed_by=request.user if request.user.is_authenticated else None,
                note=reason,
            )
            messages.success(request, "Costing unlocked.")
            return redirect("cost_sheet_detail", pk=pk)

        if action == "upload_document":
            form = OpportunityDocumentForm(request.POST, request.FILES)
            if form.is_valid():
                doc = form.save(commit=False)
                doc.opportunity = costing.opportunity
                doc.costing_header = costing
                doc.uploaded_by = request.user if request.user.is_authenticated else None
                doc.save()
                CostingAuditLog.objects.create(
                    costing=costing,
                    action="uploaded_file",
                    changed_by=request.user if request.user.is_authenticated else None,
                    note=doc.original_name or doc.file.name,
                )
                messages.success(request, "Document uploaded.")
            else:
                messages.error(request, "Please choose a file and type.")
            return redirect("cost_sheet_detail", pk=pk)

    calc = compute_costing(costing.id)
    margin_tone = "neutral"
    if calc:
        margin_percent = calc.get("margin_percent") or Decimal("0")
        if margin_percent >= Decimal("20"):
            margin_tone = "good"
        elif margin_percent >= Decimal("5"):
            margin_tone = "watch"
        else:
            margin_tone = "risk"
    grouped_lines = _group_lines(costing)
    section_total_labels = {
        "fabric": "Fabric Total",
        "sewing_trim": "Sewing Trims Total",
        "packaging_trim": "Packaging Total",
        "labels_branding": "Labels Total",
        "wash_process": "Wash / Process Total",
        "cm_labor": "CM / Labor Total",
        "logistics_compliance": "Logistics Total",
        "other": "Other Total",
    }
    category_sections = [
        {
            "key": key,
            "label": label,
            "total_label": section_total_labels.get(key, f"{label} Total"),
            "rows": grouped_lines.get(key, []),
        }
        for key, label in NEW_COSTING_CATEGORY_CHOICES
    ]

    documents = OpportunityDocument.objects.filter(
        opportunity=costing.opportunity,
        doc_type__in=["costing_pdf", "costing_excel", "costing_other"],
    ).order_by("-uploaded_at")
    audits = costing.audits.select_related("changed_by").all()[:8]
    snapshots = costing.snapshots.all()[:6]

    form = CostingHeaderForm(instance=costing)
    smv_form = CostingSMVForm(instance=_safe_costing_smv(costing))
    if "opportunity" in form.fields:
        form.fields["opportunity"].disabled = True
    if "customer" in form.fields:
        form.fields["customer"].disabled = True
    workflow = _workflow_context(costing, request.user)
    workflow_visibility = build_workflow_visibility_context(
        "costing",
        user=request.user,
        opportunity=costing.opportunity,
        costing=costing,
        quotation=costing if workflow["is_quotation"] else None,
        invoice=workflow["invoice"],
        production_order=workflow["production_order"],
        lifecycle=workflow["lifecycle"],
    )

    context = {
        "costing": costing,
        "calc": calc,
        "margin_tone": margin_tone,
        "form": form,
        "smv_form": smv_form,
        "documents": documents,
        "audits": audits,
        "snapshots": snapshots,
        "document_form": OpportunityDocumentForm(),
        "grouped_lines": grouped_lines,
        "category_sections": category_sections,
        "category_choices": NEW_COSTING_CATEGORY_CHOICES,
        "uom_choices": NEW_COSTING_UOM_CHOICES,
        "can_approve": can_approve,
        "is_locked": is_locked,
        "workflow": workflow,
        "can_view_internal_costing": can_view_costing_profit,
        **workflow_visibility,
    }
    return render(request, "crm/costing/costsheet_detail.html", context)


@require_POST
def cost_sheet_convert_to_quotation(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(CostingHeader.objects.select_related("opportunity", "customer"), pk=pk)
    if not _can_approve(request.user):
        messages.error(request, "You do not have permission to convert this costing to a quotation.")
        return redirect("cost_sheet_detail", pk=pk)

    try:
        convert_costing_to_quotation(costing, user=request.user)
        costing.quotation_status = CostingHeader.QUOTATION_STATUS_DRAFT
        costing.quotation_approved_by = None
        costing.quotation_approved_at = None
        costing.quotation_rejected_by = None
        costing.quotation_rejected_at = None
        costing.save(
            update_fields=[
                "quotation_status",
                "quotation_approved_by",
                "quotation_approved_at",
                "quotation_rejected_by",
                "quotation_rejected_at",
                "updated_at",
            ]
        )
    except CostingWorkflowError as exc:
        messages.error(request, str(exc))
        return redirect("cost_sheet_detail", pk=pk)

    messages.success(request, f"Quotation {costing.quotation_number} is ready.")
    return redirect("cost_sheet_client_quotation", pk=pk)


def cost_sheet_client_quotation(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(CostingHeader.objects.select_related("opportunity", "customer"), pk=pk)
    if costing.status != "approved":
        messages.error(request, "Approve the costing before viewing the client quotation.")
        return redirect("cost_sheet_detail", pk=pk)
    if not costing.quotation_number:
        messages.error(request, "Convert this approved costing to a quotation first.")
        return redirect("cost_sheet_detail", pk=pk)

    try:
        context = _quotation_context(costing, request.user)
    except CostingWorkflowError as exc:
        messages.error(request, str(exc))
        return redirect("cost_sheet_detail", pk=pk)
    workflow = _workflow_context(costing, request.user)
    context.update(
        build_workflow_visibility_context(
            "quotation",
            user=request.user,
            opportunity=costing.opportunity,
            costing=costing,
            quotation=costing,
            invoice=workflow["invoice"],
            production_order=workflow["production_order"],
            lifecycle=workflow["lifecycle"],
        )
    )

    return render(request, "crm/costing/quotation_client.html", context)


@require_POST
def cost_sheet_quotation_approve(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(CostingHeader.objects.select_related("opportunity", "customer"), pk=pk)
    if not _can_approve(request.user):
        messages.error(request, "You do not have permission to approve quotations.")
        return redirect("cost_sheet_client_quotation", pk=pk)
    if costing.status != "approved" or not costing.quotation_number:
        messages.error(request, "Convert an approved costing to a quotation before approval.")
        return redirect("cost_sheet_detail", pk=pk)

    costing.quotation_status = CostingHeader.QUOTATION_STATUS_APPROVED
    costing.quotation_approved_by = _user_or_none(request.user)
    costing.quotation_approved_at = timezone.now()
    costing.quotation_rejected_by = None
    costing.quotation_rejected_at = None
    costing.save(
        update_fields=[
            "quotation_status",
            "quotation_approved_by",
            "quotation_approved_at",
            "quotation_rejected_by",
            "quotation_rejected_at",
            "updated_at",
        ]
    )
    CostingAuditLog.objects.create(
        costing=costing,
        action="quotation_approved",
        changed_by=_user_or_none(request.user),
    )
    messages.success(request, f"Quotation {costing.quotation_number} approved.")
    return redirect("cost_sheet_client_quotation", pk=pk)


@require_POST
def cost_sheet_quotation_reject(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(CostingHeader.objects.select_related("opportunity", "customer"), pk=pk)
    if not _can_approve(request.user):
        messages.error(request, "You do not have permission to reject quotations.")
        return redirect("cost_sheet_client_quotation", pk=pk)
    if costing.status != "approved" or not costing.quotation_number:
        messages.error(request, "Convert an approved costing to a quotation before rejection.")
        return redirect("cost_sheet_detail", pk=pk)

    costing.quotation_status = CostingHeader.QUOTATION_STATUS_REJECTED
    costing.quotation_rejected_by = _user_or_none(request.user)
    costing.quotation_rejected_at = timezone.now()
    costing.quotation_approved_by = None
    costing.quotation_approved_at = None
    costing.save(
        update_fields=[
            "quotation_status",
            "quotation_rejected_by",
            "quotation_rejected_at",
            "quotation_approved_by",
            "quotation_approved_at",
            "updated_at",
        ]
    )
    CostingAuditLog.objects.create(
        costing=costing,
        action="quotation_rejected",
        changed_by=_user_or_none(request.user),
    )
    messages.success(request, f"Quotation {costing.quotation_number} rejected.")
    return redirect("cost_sheet_client_quotation", pk=pk)


def cost_sheet_quotation_pdf(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(CostingHeader.objects.select_related("opportunity", "customer"), pk=pk)
    if costing.status != "approved" or not costing.quotation_number:
        messages.error(request, "Convert this approved costing to a quotation before downloading the quotation PDF.")
        return redirect("cost_sheet_detail", pk=pk)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError:
        messages.error(request, "PDF export is unavailable. Please install ReportLab.")
        return redirect("cost_sheet_client_quotation", pk=pk)

    try:
        context = _quotation_context(costing, request.user)
        amounts = context["amounts"]
        company = context["company"]
        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter, pageCompression=0)
        width, height = letter
        left = 46
        right = width - 46
        y = height - 46

        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.rect(0, height - 104, width, 104, fill=1, stroke=0)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 18)
        pdf.drawString(left, height - 58, company["name"])
        pdf.setFont("Helvetica", 9)
        pdf.drawString(left, height - 76, "Professional apparel manufacturing quotation")
        pdf.setFont("Helvetica-Bold", 18)
        pdf.drawRightString(right, height - 58, "QUOTATION")
        pdf.setFont("Helvetica", 9)
        pdf.drawRightString(right, height - 76, f"Quote # {costing.quotation_number}")

        y = height - 132
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(left, y, "From")
        pdf.drawString(width / 2 + 12, y, "Quotation Details")
        y -= 15

        pdf.setFont("Helvetica", 9)
        company_lines = [
            company.get("name", ""),
            company.get("address", ""),
            " | ".join(part for part in [company.get("phone", ""), company.get("email", "")] if part),
            company.get("website", ""),
        ]
        detail_lines = [
            f"Quote date: {costing.quoted_at:%Y-%m-%d}" if costing.quoted_at else f"Quote date: {timezone.localdate():%Y-%m-%d}",
            f"Opportunity: {costing.opportunity.opportunity_id}",
            f"Currency: {costing.currency}",
            f"Reference: {costing.quotation_number}",
        ]
        row_y = y
        for line in [line for line in company_lines if line]:
            pdf.drawString(left, row_y, line[:84])
            row_y -= 12
        row_y = y
        for line in detail_lines:
            pdf.drawString(width / 2 + 12, row_y, line[:70])
            row_y -= 12

        y -= 78
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(left, y, "Customer")
        y -= 15
        pdf.setFont("Helvetica", 9)
        customer = costing.customer
        customer_lines = [
            (getattr(customer, "account_brand", "") or getattr(customer, "contact_name", "") or "Customer") if customer else "Customer",
            getattr(customer, "contact_name", "") if customer else "",
            getattr(customer, "email", "") if customer else "",
            getattr(customer, "phone", "") if customer else "",
        ]
        for line in [line for line in customer_lines if line]:
            pdf.drawString(left, y, line[:92])
            y -= 12

        y -= 16
        pdf.setFillColor(colors.HexColor("#f3f4f6"))
        pdf.rect(left, y - 98, right - left, 118, fill=1, stroke=0)
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(left + 10, y, "Item")
        pdf.drawRightString(right - 190, y, "Qty")
        pdf.drawRightString(right - 95, y, "Unit Price")
        pdf.drawRightString(right - 10, y, "Amount")
        y -= 20
        pdf.setFont("Helvetica", 9)
        description = costing.style_name or costing.style_code or costing.get_product_type_display()
        pdf.drawString(left + 10, y, description[:58])
        pdf.drawRightString(right - 190, y, f"{amounts['quantity']}")
        pdf.drawRightString(right - 95, y, f"{costing.currency} {amounts['unit_price']:,.2f}")
        pdf.drawRightString(right - 10, y, f"{costing.currency} {amounts['order_total']:,.2f}")
        y -= 18
        specs = " | ".join(
            part
            for part in [
                f"Fabric: {costing.fabric_type}" if costing.fabric_type else "",
                f"GSM: {costing.fabric_gsm}" if costing.fabric_gsm else "",
                f"Composition: {costing.fabric_composition}" if costing.fabric_composition else "",
                f"Packaging: {costing.packaging_type}" if costing.packaging_type else "",
            ]
            if part
        )
        pdf.drawString(left + 10, y, (specs or "Apparel production quotation")[:110])
        y -= 38
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawRightString(right - 10, y, f"Total: {costing.currency} {amounts['order_total']:,.2f}")

        y -= 44
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(left, y, "Terms and Conditions")
        y -= 14
        pdf.setFont("Helvetica", 8.5)
        for paragraph in DEFAULT_QUOTATION_TERMS.splitlines():
            if not paragraph.strip():
                y -= 5
                continue
            for line in _pdf_lines(pdf, paragraph, right - left, "Helvetica", 8.5):
                if y < 48:
                    pdf.showPage()
                    y = height - 46
                    pdf.setFont("Helvetica", 8.5)
                pdf.drawString(left, y, line)
                y -= 11

        pdf.showPage()
        pdf.save()
        data = buffer.getvalue()
    except CostingWorkflowError as exc:
        messages.error(request, str(exc))
        return redirect("cost_sheet_detail", pk=pk)
    except Exception:
        logger.exception("Failed to generate client quotation PDF", extra={"costing_header": costing.pk})
        messages.error(request, "Could not generate the quotation PDF. Please try again.")
        return redirect("cost_sheet_client_quotation", pk=pk)

    filename = f"quotation_{costing.quotation_number or costing.pk}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(data)
    return response


@require_POST
def cost_sheet_convert_to_invoice(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(CostingHeader.objects.select_related("opportunity", "customer"), pk=pk)
    if not _can_convert_to_invoice(request.user):
        messages.error(request, "Only invoice managers can convert quotations to invoices.")
        return redirect("cost_sheet_detail", pk=pk)
    if costing.quotation_status != CostingHeader.QUOTATION_STATUS_APPROVED:
        messages.error(request, "Approve the quotation before creating an invoice.")
        return redirect("cost_sheet_client_quotation", pk=pk)

    try:
        invoice, created = create_invoice_from_costing(costing, user=request.user)
    except CostingWorkflowError as exc:
        messages.error(request, str(exc))
        return redirect("cost_sheet_detail", pk=pk)

    if created:
        messages.success(request, f"Invoice {invoice.invoice_number} created from quotation.")
    else:
        messages.info(request, f"Invoice {invoice.invoice_number} already exists for this quotation.")
    return redirect("invoice_view", pk=invoice.pk)


def cost_sheet_duplicate(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(
        CostingHeader.objects.select_related("opportunity", "customer").prefetch_related("line_items"),
        pk=pk,
    )
    new_costing = CostingHeader.objects.create(
        opportunity=costing.opportunity,
        customer=costing.customer,
        buyer=costing.buyer,
        brand=costing.brand,
        style_name=costing.style_name,
        style_code=costing.style_code,
        product_type=costing.product_type,
        gender=costing.gender,
        size_range=costing.size_range,
        season=costing.season,
        factory_location=costing.factory_location,
        order_quantity=costing.order_quantity,
        moq=costing.moq,
        costing_date=costing.costing_date,
        currency=costing.currency,
        exchange_rate=costing.exchange_rate,
        finance_percent_fabric=costing.finance_percent_fabric,
        finance_percent_trims=costing.finance_percent_trims,
        commission_percent=costing.commission_percent,
        target_margin_percent=costing.target_margin_percent,
        manual_fob_per_piece=costing.manual_fob_per_piece,
        shipping_cost=costing.shipping_cost,
        merchandiser=costing.merchandiser,
        fabric_type=costing.fabric_type,
        fabric_gsm=costing.fabric_gsm,
        fabric_composition=costing.fabric_composition,
        wash_type=costing.wash_type,
        print_type=costing.print_type,
        embroidery=costing.embroidery,
        label_type=costing.label_type,
        packaging_type=costing.packaging_type,
        special_trims=costing.special_trims,
        fit_remarks=costing.fit_remarks,
        notes=costing.notes,
        status="draft",
    )

    for line in costing.line_items.all():
        line.pk = None
        line.costing = new_costing
        line.save()

    smv = _safe_costing_smv(costing)
    if smv:
        CostingSMV.objects.create(
            costing=new_costing,
            machine_smv=smv.machine_smv,
            finishing_smv=smv.finishing_smv,
            cpm=smv.cpm,
            efficiency_costing=smv.efficiency_costing,
            efficiency_planned=smv.efficiency_planned,
        )

    CostingAuditLog.objects.create(
        costing=new_costing,
        action="created",
        changed_by=request.user if request.user.is_authenticated else None,
        note=f"Duplicated from COST-{costing.pk}",
    )
    messages.success(request, "Costing duplicated. You are now editing the new version.")
    return redirect("cost_sheet_detail", pk=new_costing.pk)


def _save_export_document(costing, filename, data, doc_type, user):
    try:
        OpportunityDocument.objects.create(
            opportunity=costing.opportunity,
            costing_header=costing,
            file=ContentFile(data, name=filename),
            original_name=filename,
            doc_type=doc_type,
            uploaded_by=user if user and user.is_authenticated else None,
        )
        CostingAuditLog.objects.create(
            costing=costing,
            action="exported",
            changed_by=user if user and user.is_authenticated else None,
            note=filename,
        )
    except Exception:
        logger.exception("Failed to save costing export document", extra={"costing_header": costing.pk})


def cost_sheet_export_pdf(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(
        CostingHeader.objects.select_related("opportunity", "customer").prefetch_related("line_items"),
        pk=pk,
    )

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError:
        messages.error(request, "PDF export is unavailable. Please install ReportLab.")
        return redirect("cost_sheet_detail", pk=pk)

    try:
        calc = compute_costing(costing.id)
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4, pageCompression=0)
        width, height = A4
        left = 36
        right = width - 36
        black = colors.HexColor("#111111")
        gold = colors.HexColor("#c89b3c")
        pale = colors.HexColor("#f6f6f4")
        border = colors.HexColor("#e5e7eb")
        muted = colors.HexColor("#6b7280")
        currency = _costing_currency(costing)

        def money(value):
            value = value or Decimal("0")
            return f"{currency} {Decimal(value).quantize(Decimal('0.01')):,.2f}"

        def text(value, fallback="-"):
            value = value if value not in (None, "") else fallback
            return str(value)

        def draw_logo(x, y):
            try:
                from django.contrib.staticfiles import finders

                logo_path = finders.find("img/image.png")
            except Exception:
                logo_path = None
            if logo_path:
                p.drawImage(logo_path, x, y - 34, width=42, height=42, preserveAspectRatio=True, mask="auto")
                return
            p.setStrokeColor(gold)
            p.setLineWidth(1.2)
            p.circle(x + 18, y - 14, 20, fill=0, stroke=1)
            p.setFillColor(colors.white)
            p.setFont("Helvetica-Bold", 10)
            p.drawCentredString(x + 18, y - 18, "IAH")

        def draw_page_brand_header():
            p.setFillColor(black)
            p.rect(0, height - 102, width, 102, fill=1, stroke=0)
            p.setFillColor(gold)
            p.rect(0, height - 102, width, 5, fill=1, stroke=0)
            p.rect(right - 148, height - 66, 148, 1.4, fill=1, stroke=0)
            draw_logo(left, height - 25)
            p.setFillColor(colors.white)
            p.setFont("Helvetica-Bold", 15)
            p.drawString(left + 52, height - 38, "Iconic Apparel House")
            p.setFont("Helvetica", 8.8)
            p.drawString(left + 52, height - 54, "Premium apparel sourcing, development, and production")
            p.setFont("Helvetica-Bold", 10)
            p.drawRightString(right, height - 38, f"COST-{costing.pk}")
            p.setFont("Helvetica", 8.5)
            p.drawRightString(right, height - 54, timezone.localdate().strftime("%Y-%m-%d"))

        def draw_table_header(y_pos):
            p.setFillColor(gold)
            p.rect(left, y_pos - 22, right - left, 22, fill=1, stroke=0)
            p.setFillColor(black)
            p.setFont("Helvetica-Bold", 8.5)
            p.drawString(left + 8, y_pos - 14, "SL")
            p.drawString(left + 40, y_pos - 14, "Description")
            p.drawString(left + 220, y_pos - 14, "Calculation")
            p.drawRightString(right - 8, y_pos - 14, "Amount")
            return y_pos - 22

        def draw_wrapped(lines, x, y_pos, max_width, font="Helvetica", size=8.2, line_gap=10):
            p.setFont(font, size)
            current_y = y_pos
            for line in lines:
                for wrapped in _pdf_lines(p, line, max_width, font, size):
                    p.drawString(x, current_y, wrapped)
                    current_y -= line_gap
            return current_y

        draw_page_brand_header()
        y = height - 132

        project_name = costing.style_name or costing.style_code or (
            costing.opportunity.opportunity_id if costing.opportunity_id else f"Costing {costing.pk}"
        )
        item_name = costing.get_product_type_display()
        buyer_name = costing.buyer or costing.brand or (
            getattr(costing.customer, "account_brand", "") if costing.customer else ""
        ) or "Buyer not set"
        prepared_by = costing.merchandiser or (
            request.user.get_full_name() or request.user.get_username()
            if request.user.is_authenticated
            else "Iconic Team"
        )

        p.setFillColor(black)
        p.setFont("Helvetica-Bold", 24)
        p.drawCentredString(width / 2, y, "COSTING SHEET")
        y -= 26
        p.setFillColor(gold)
        p.setFont("Helvetica-Bold", 13)
        p.drawCentredString(width / 2, y, text(project_name, "Project"))
        y -= 18
        p.setFillColor(black)
        p.setFont("Helvetica-Bold", 10)
        p.drawCentredString(width / 2, y, f"{costing.order_quantity or 0} PCS {item_name}".upper())
        y -= 24

        box_width = (right - left - 12) / 2
        p.setFillColor(pale)
        p.roundRect(left, y - 42, box_width, 42, 6, fill=1, stroke=0)
        p.roundRect(left + box_width + 12, y - 42, box_width, 42, 6, fill=1, stroke=0)
        p.setFillColor(muted)
        p.setFont("Helvetica-Bold", 8)
        p.drawString(left + 10, y - 15, "BUYER NAME")
        p.drawString(left + box_width + 22, y - 15, "DATE")
        p.setFillColor(black)
        p.setFont("Helvetica-Bold", 11)
        p.drawString(left + 10, y - 31, text(buyer_name))
        p.drawString(left + box_width + 22, y - 31, (costing.costing_date or timezone.localdate()).strftime("%Y-%m-%d"))
        y -= 62

        rows = []
        for row in calc["line_rows"]:
            row_amount = (row["cost_per_piece"] or Decimal("0")) * Decimal(calc["order_quantity"] or 0)
            calc_bits = [
                f"UOM: {row['uom']}",
                f"Cons: {row['consumption_value']}",
                f"Rate: {money(row['unit_price'])}",
            ]
            if row.get("freight"):
                calc_bits.append(f"Freight: {money(row['freight'])}")
            if row.get("wastage_percent"):
                calc_bits.append(f"Wastage: {row['wastage_percent']}%")
            rows.append(
                {
                    "description": row["item_name"] or row["category"].replace("_", " ").title(),
                    "calculation": " | ".join(calc_bits),
                    "amount": row_amount,
                }
            )

        rows.append(
            {
                "description": "Shipping Cost",
                "calculation": "Order-level shipping cost",
                "amount": calc.get("shipping_cost_order", Decimal("0")),
            }
        )

        y = draw_table_header(y)
        for index, row in enumerate(rows, start=1):
            desc_lines = _pdf_lines(p, row["description"], 162, "Helvetica", 8.2)
            calc_lines = _pdf_lines(p, row["calculation"], 200, "Helvetica", 8.2)
            row_height = max(30, (max(len(desc_lines), len(calc_lines), 1) * 10) + 14)
            if y - row_height < 125:
                p.showPage()
                draw_page_brand_header()
                y = draw_table_header(height - 132)
            p.setFillColor(colors.white if index % 2 else colors.HexColor("#fcfcfd"))
            p.rect(left, y - row_height, right - left, row_height, fill=1, stroke=0)
            p.setStrokeColor(border)
            p.line(left, y - row_height, right, y - row_height)
            p.setFillColor(black)
            p.setFont("Helvetica", 8.4)
            p.drawString(left + 8, y - 18, str(index))
            draw_wrapped(desc_lines, left + 40, y - 16, 162)
            draw_wrapped(calc_lines, left + 220, y - 16, 200)
            p.setFont("Helvetica-Bold", 8.6)
            p.drawRightString(right - 8, y - 18, money(row["amount"]))
            y -= row_height

        p.setFillColor(black)
        p.rect(left, y - 28, right - left, 28, fill=1, stroke=0)
        p.setFillColor(colors.white)
        p.setFont("Helvetica-Bold", 11)
        p.drawString(left + 8, y - 18, "Total Amount")
        p.drawRightString(right - 8, y - 18, money(calc["total_cost_order"]))
        y -= 48

        if y < 168:
            p.showPage()
            draw_page_brand_header()
            y = height - 132

        p.setFillColor(black)
        p.roundRect(left, y - 112, right - left, 112, 8, fill=1, stroke=0)
        p.setFillColor(gold)
        p.rect(left, y - 4, right - left, 4, fill=1, stroke=0)
        p.setFillColor(colors.white)
        p.setFont("Helvetica-Bold", 12)
        p.drawString(left + 12, y - 18, "Summary")
        summary_rows = [
            ("Total Amount", money(calc["total_cost_order"])),
            ("Total Pieces", text(costing.order_quantity or 0)),
            ("Buyer Name", text(buyer_name)),
            ("Project", text(project_name)),
            ("Item", text(item_name)),
            ("Prepared By", text(prepared_by)),
        ]
        summary_y = y - 40
        col_x = [left + 12, left + 286]
        for idx, (label, value) in enumerate(summary_rows):
            x = col_x[idx % 2]
            if idx and idx % 2 == 0:
                summary_y -= 24
            p.setFillColor(gold)
            p.setFont("Helvetica-Bold", 7.8)
            p.drawString(x, summary_y, label.upper())
            p.setFillColor(colors.white)
            p.setFont("Helvetica-Bold", 9.2)
            p.drawString(x, summary_y - 13, value[:42])

        footer_y = 62
        p.setFillColor(black)
        p.rect(0, 0, width, footer_y + 30, fill=1, stroke=0)
        p.setFillColor(gold)
        p.rect(0, footer_y + 28, width, 3, fill=1, stroke=0)
        p.setFillColor(colors.white)
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(width / 2, footer_y + 10, "Thank You!")
        p.setFillColor(gold)
        p.setFont("Helvetica", 10)
        p.drawCentredString(width / 2, footer_y - 5, "For Your Business")

        p.save()
        pdf_bytes = buffer.getvalue()
    except Exception:
        logger.exception("Failed to generate costing PDF", extra={"costing_header": costing.pk})
        messages.error(request, "Could not generate the PDF. Please try again.")
        return redirect("cost_sheet_detail", pk=pk)

    filename = f"costing_{costing.opportunity.opportunity_id}.pdf"
    _save_export_document(costing, filename, pdf_bytes, "costing_pdf", request.user)

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write(pdf_bytes)
    return resp


def cost_sheet_export_excel(request, pk):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    costing = get_object_or_404(
        CostingHeader.objects.select_related("opportunity", "customer").prefetch_related("line_items"),
        pk=pk,
    )

    try:
        from openpyxl import Workbook
    except Exception:
        messages.error(request, "Excel export is unavailable. Please install openpyxl.")
        return redirect("cost_sheet_detail", pk=pk)

    try:
        calc = compute_costing(costing.id)
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary.append(["Customer", (costing.customer.account_brand if costing.customer else "") or "Not set"])
        ws_summary.append(["Opportunity", costing.opportunity.opportunity_id])
        ws_summary.append(["Style name", costing.style_name or "-"])
        ws_summary.append(["Style code", costing.style_code or "-"])
        ws_summary.append(["Product type", costing.get_product_type_display()])
        ws_summary.append(["Quantity", costing.order_quantity])
        ws_summary.append(["Factory location", costing.get_factory_location_display()])
        ws_summary.append(["Status", costing.get_status_display()])
        ws_summary.append(["Currency", costing.currency])
        exchange_rate_label = f"Per 1 {_costing_currency(costing)}" if _costing_currency(costing) != "BDT" else ""
        ws_summary.append(["Exchange rate", costing.exchange_rate or "", exchange_rate_label])

        ws_summary.append([])
        ws_summary.append(["Total cost per piece", _format_costing_money(costing, calc["display"]["total_cost_per_piece"])])
        ws_summary.append(["FOB per piece", _format_costing_money(costing, calc["display"]["fob_per_piece"])])
        ws_summary.append(["Profit per piece", _format_costing_money(costing, calc["display"]["profit_per_piece"])])
        ws_summary.append(["Margin %", float(calc["display"]["margin_percent"])])
        ws_summary.append(["Total cost order", _format_costing_money(costing, calc["display"]["total_cost_order"])])
        ws_summary.append(["Total sales order", _format_costing_money(costing, calc["display"]["total_sales_order"])])
        ws_summary.append(["Total profit order", _format_costing_money(costing, calc["display"]["total_profit_order"])])

        ws_lines = wb.create_sheet("Line items")
        currency = _costing_currency(costing)
        ws_lines.append([
            "Category",
            "Item",
            "UOM",
            f"Unit price ({currency})",
            f"Freight ({currency})",
            "Consumption",
            "Wastage %",
            "Denominator",
            f"Cost per piece ({currency})",
        ])
        for row in calc["line_rows"]:
            ws_lines.append([
                row["category"],
                row["item_name"],
                row["uom"],
                float(row["unit_price"]),
                float(row["freight"]),
                float(row["consumption_value"]),
                float(row["wastage_percent"]),
                float(row["denominator_value"] or 0),
                float(row["cost_per_piece"]),
            ])

        output = io.BytesIO()
        wb.save(output)
        data = output.getvalue()
    except Exception:
        logger.exception("Failed to generate costing Excel", extra={"costing_header": costing.pk})
        messages.error(request, "Could not generate the Excel file. Please try again.")
        return redirect("cost_sheet_detail", pk=pk)

    filename = f"costing_{costing.opportunity.opportunity_id}.xlsx"
    _save_export_document(costing, filename, data, "costing_excel", request.user)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write(data)
    return resp


def cost_sheet_dashboard(request):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    can_view_costing_profit = can_view_internal_costing(request.user)
    qs = CostingHeader.objects.select_related("customer", "opportunity").order_by("-updated_at")

    approved_only = (request.GET.get("approved") or "").strip() == "1"
    customer_id = (request.GET.get("customer") or "").strip()
    product_type = (request.GET.get("product_type") or "").strip()
    factory_location = (request.GET.get("factory_location") or "").strip()
    currency = (request.GET.get("currency") or "").strip().upper()
    start_date = (request.GET.get("start") or "").strip()
    end_date = (request.GET.get("end") or "").strip()

    if approved_only:
        qs = qs.filter(status="approved")
    if customer_id:
        qs = qs.filter(customer_id=customer_id)
    if product_type:
        qs = qs.filter(product_type=product_type)
    if factory_location:
        qs = qs.filter(factory_location=factory_location)
    if currency:
        qs = qs.filter(currency=currency)
    if start_date:
        qs = qs.filter(updated_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(updated_at__date__lte=end_date)

    rows = []
    for cost in qs:
        calc = compute_costing(cost.id)
        if calc:
            rows.append(calc)

    top_profit = sorted(rows, key=lambda r: r["total_profit_order"], reverse=True)[:10]
    lowest_margin = sorted(rows, key=lambda r: r["margin_percent"])[:10]

    breakdown_totals = defaultdict(Decimal)
    for row in rows:
        breakdown_totals["fabric"] += row["fabric_base"]
        breakdown_totals["trims"] += row["trims_base"]
        breakdown_totals["labor"] += row["labor_cost_per_piece"]
        breakdown_totals["other"] += row["other_base"]

    trend = defaultdict(list)
    for row in rows:
        key = row["costing"].updated_at.strftime("%Y-%m")
        trend[key].append(row["total_cost_per_piece"])

    trend_labels = sorted(trend.keys())
    trend_values = [float(sum(trend[k]) / len(trend[k])) for k in trend_labels]

    customer_summary = defaultdict(list)
    for row in rows:
        customer = row["costing"].customer
        if not customer:
            continue
        customer_summary[customer].append(row)

    customer_rows = []
    for customer, items in customer_summary.items():
        total_qty = sum(r["order_quantity"] for r in items)
        avg_margin = sum(r["margin_percent"] for r in items) / Decimal(len(items)) if items else Decimal("0")
        customer_rows.append(
            {
                "customer": customer,
                "total_qty": total_qty,
                "avg_margin": avg_margin,
            }
        )

    context = {
        "top_profit": top_profit,
        "lowest_margin": lowest_margin,
        "breakdown_totals": breakdown_totals,
        "trend_labels": trend_labels,
        "trend_values": trend_values,
        "customer_rows": customer_rows,
        "filters": {
            "approved": approved_only,
            "customer": customer_id,
            "product_type": product_type,
            "factory_location": factory_location,
            "currency": currency,
            "start": start_date,
            "end": end_date,
        },
        "product_types": Opportunity.PRODUCT_TYPE_CHOICES,
        "currencies": NEW_COSTING_CURRENCY_CHOICES,
        "factory_locations": [
            ("bd", "Bangladesh"),
            ("ca", "Canada"),
            ("other", "Other"),
        ],
        "customers": list({row["costing"].customer for row in rows if row["costing"].customer}),
        "can_view_internal_costing": can_view_costing_profit,
    }
    return render(request, "crm/costing/costing_dashboard.html", context)


def cost_sheet_reports(request):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    can_view_costing_profit = can_view_internal_costing(request.user)
    qs = CostingHeader.objects.select_related("customer", "opportunity").order_by("-updated_at")
    export = (request.GET.get("export") or "").strip()

    rows = []
    for cost in qs:
        calc = compute_costing(cost.id)
        if calc:
            rows.append(calc)

    if export:
        output = io.StringIO()
        if export == "list":
            output.write("Opportunity,Customer,Style,Currency,Qty,Cost per piece,FOB per piece,Margin %\n")
            for row in rows:
                cost = row["costing"]
                output.write(
                    f"{cost.opportunity.opportunity_id},{(cost.customer.account_brand if cost.customer else '')},{cost.style_name},{_costing_currency(cost)},{row['order_quantity']},{row['total_cost_per_piece']},{row['fob_per_piece']},{row['margin_percent']}\n"
                )
        elif export == "margin":
            output.write("Opportunity,Style,Currency,Margin %,Total profit\n")
            for row in rows:
                cost = row["costing"]
                output.write(
                    f"{cost.opportunity.opportunity_id},{cost.style_name},{_costing_currency(cost)},{row['margin_percent']},{row['total_profit_order']}\n"
                )
        elif export == "finance":
            output.write("Month,Currency,Fabric finance,Trim finance\n")
            month_totals = defaultdict(lambda: {"fabric": Decimal("0"), "trims": Decimal("0")})
            for row in rows:
                key = (row["costing"].updated_at.strftime("%Y-%m"), _costing_currency(row["costing"]))
                month_totals[key]["fabric"] += row["fabric_finance"] * Decimal(row["order_quantity"])
                month_totals[key]["trims"] += row["trims_finance"] * Decimal(row["order_quantity"])
            for key in sorted(month_totals.keys()):
                month, currency = key
                output.write(f"{month},{currency},{month_totals[key]['fabric']},{month_totals[key]['trims']}\n")
        else:
            output.write("Style,Currency,Old cost per piece,New cost per piece,Delta\n")
            by_style = defaultdict(list)
            for row in rows:
                by_style[(row["costing"].style_code or row["costing"].style_name, _costing_currency(row["costing"]))].append(row)
            for style, items in by_style.items():
                if len(items) < 2:
                    continue
                items_sorted = sorted(items, key=lambda r: r["costing"].updated_at)
                old = items_sorted[0]["total_cost_per_piece"]
                new = items_sorted[-1]["total_cost_per_piece"]
                style_name, currency = style
                output.write(f"{style_name},{currency},{old},{new},{new - old}\n")

        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="costing_{export}_report.csv"'
        resp.write(output.getvalue())
        return resp

    context = {
        "rows": rows,
        "can_view_internal_costing": can_view_costing_profit,
    }
    return render(request, "crm/costing/costing_reports.html", context)


def cost_sheet_guide(request):
    denied = _deny_without_internal_costing(request)
    if denied:
        return denied
    return render(request, "crm/costing/costing_guide.html")
