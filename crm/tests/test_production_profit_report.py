from decimal import Decimal

from django.contrib.auth import get_user_model
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone

from crm.models import (
    AccountingEntry,
    Customer,
    ExchangeRate,
    Invoice,
    Lead,
    Opportunity,
    ProductionOrder,
    QuickCosting,
)
from crm.services.production_profit import build_production_profit_report


class ProductionProfitReportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_superuser(
            username="production-profit-admin",
            email="production-profit@example.com",
            password="pass",
        )

    def setUp(self):
        self.client.force_login(self.user)
        self.today = timezone.localdate()

    def report(self, **kwargs):
        return build_production_profit_report(
            year=self.today.year,
            month=self.today.month,
            **kwargs,
        )

    def canada_order(
        self,
        *,
        code,
        cost_bdt=None,
        sewing_cost_bdt=None,
        quantity=100,
        **extra,
    ):
        return ProductionOrder.objects.create(
            title=f"Canada order {code}",
            order_code=code,
            order_type="fob",
            factory_location="bd",
            qty_total=quantity,
            production_total_cost_bdt=cost_bdt,
            production_sewing_cost_bdt=sewing_cost_bdt,
            **extra,
        )

    def local_order(self, *, code, charge=None, cost=None, extra=Decimal("0")):
        quick = QuickCosting.objects.create(
            buyer_name=f"Local buyer {code}",
            project_name=f"Local project {code}",
            product_type="Other",
            pricing_type=QuickCosting.PRICING_CMT,
            currency="BDT",
            quantity=100,
            sewing_charge_per_piece_bdt=charge,
            sewing_cost_per_piece_bdt=cost,
            extra_local_cost_bdt=extra,
            status=QuickCosting.STATUS_APPROVED,
            approved_at=timezone.now(),
        )
        return ProductionOrder.objects.create(
            title=f"Local order {code}",
            order_code=code,
            order_type="sewing_charge",
            factory_location="bd",
            qty_total=100,
            sewing_charge_per_piece_bdt=charge,
            sewing_cost_per_piece_bdt=cost,
            extra_local_cost_bdt=extra,
            source_quick_costing=quick,
        )

    def invoice(
        self,
        order=None,
        *,
        number,
        amount,
        currency,
        market,
        region,
        invoice_type="bulk",
        issue_date=None,
        paid_amount=Decimal("0"),
        customer=None,
        quick_costing=None,
        costing_header=None,
        shipping_amount=Decimal("0"),
        sewing_charge=Decimal("0"),
        other_internal_cost=Decimal("0"),
        notes="",
    ):
        values = {
            "invoice_number": number,
            "order": order,
            "issue_date": issue_date or self.today,
            "currency": currency,
            "invoice_market": market,
            "invoice_region": region,
            "invoice_type": invoice_type,
            "subtotal": amount,
            "total_amount": amount,
            "paid_amount": paid_amount,
            "customer": customer,
            "quick_costing": quick_costing,
            "costing_header": costing_header,
            "shipping_amount": shipping_amount,
            "sewing_charge": sewing_charge,
            "other_internal_cost": other_internal_cost,
            "notes": notes,
        }
        if order and order.source_quick_costing_id:
            values["quick_costing"] = order.source_quick_costing
        return Invoice.objects.create(**values)

    def sample_invoice(self, order=None, *, number="INV-SAMPLE-001", **extra):
        defaults = {
            "amount": Decimal("150"),
            "currency": "CAD",
            "market": "north_america",
            "region": "CA",
            "invoice_type": "sample",
        }
        defaults.update(extra)
        return self.invoice(order, number=number, **defaults)

    def test_canada_invoice_revenue_and_bdt_cost_convert_to_cad(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.canada_order(code="POCANADA001", cost_bdt=Decimal("85000"))
        self.invoice(
            order,
            number="INV-PROFIT-CA",
            amount=Decimal("2500"),
            currency="CAD",
            market="north_america",
            region="CA",
        )

        report = self.report()
        row = next(item for item in report["rows"] if item["production_order_id"] == order.pk)

        self.assertEqual(row["classification"], "canada_export")
        self.assertEqual(row["revenue_amount"], Decimal("2500.00"))
        self.assertEqual(row["cost_bdt"], Decimal("85000.00"))
        self.assertEqual(row["cost_cad"], Decimal("1000.00"))
        self.assertEqual(row["profit"], Decimal("1500.00"))
        self.assertEqual(row["margin_pct"], Decimal("60.00"))
        self.assertEqual(report["canada_export"]["profit"], Decimal("1500.00"))

    def test_bangladesh_local_revenue_cost_profit_and_margin_are_bdt(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.local_order(
            code="POLOCAL001",
            charge=Decimal("120"),
            cost=Decimal("70"),
            extra=Decimal("500"),
        )
        self.invoice(
            order,
            number="INV-PROFIT-BD",
            amount=Decimal("12000"),
            currency="BDT",
            market="bangladesh",
            region="BD",
            invoice_type="sewing_charge",
        )

        report = self.report()
        row = next(item for item in report["rows"] if item["production_order_id"] == order.pk)

        self.assertEqual(row["classification"], "bangladesh_local")
        self.assertEqual(row["revenue_currency"], "BDT")
        self.assertEqual(row["revenue_amount"], Decimal("12000.00"))
        self.assertEqual(row["cost_bdt"], Decimal("7500.00"))
        self.assertEqual(row["profit"], Decimal("4500.00"))
        self.assertEqual(row["margin_pct"], Decimal("37.50"))
        self.assertEqual(report["bangladesh_local_sewing"]["profit"], Decimal("4500.00"))

    def test_canada_sewing_charge_invoice_uses_sewing_cost_only(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.canada_order(
            code="POSEWINGCA001",
            cost_bdt=Decimal("85000"),
            sewing_cost_bdt=Decimal("42500"),
        )
        self.invoice(
            order,
            number="INV-PROFIT-SEW-CA",
            amount=Decimal("1000"),
            currency="CAD",
            market="north_america",
            region="CA",
            invoice_type="sewing_charge",
        )

        summary = self.report()["canada_export_sewing"]

        self.assertEqual(summary["revenue"], Decimal("1000.00"))
        self.assertEqual(summary["cost_bdt"], Decimal("42500.00"))
        self.assertEqual(summary["cost_cad"], Decimal("500.00"))
        self.assertEqual(summary["profit"], Decimal("500.00"))
        self.assertEqual(summary["margin_pct"], Decimal("50.00"))

    def test_canada_sewing_subtype_uses_linked_accounting_revenue(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.canada_order(
            code="POSEWINGSUBTYPE",
            cost_bdt=Decimal("85000"),
            sewing_cost_bdt=Decimal("17000"),
        )
        self.invoice(
            order,
            number="INV-PROFIT-BULK-WITH-SEWING",
            amount=Decimal("1000"),
            currency="CAD",
            market="north_america",
            region="CA",
            invoice_type="bulk",
        )
        AccountingEntry.objects.create(
            date=self.today,
            side="CA",
            direction="IN",
            main_type="REVENUE",
            sub_type="Swing",
            production_order=order,
            currency="CAD",
            amount_original=Decimal("400"),
            rate_to_bdt=Decimal("85"),
        )

        report = self.report()
        row = next(item for item in report["rows"] if item["production_order_id"] == order.pk)
        summary = report["canada_export_sewing"]

        self.assertEqual(row["revenue_amount"], Decimal("1000.00"))
        self.assertEqual(row["sewing_charge_amount"], Decimal("400.00"))
        self.assertEqual(row["sewing_charge_source"], "CA accounting sewing revenue")
        self.assertEqual(summary["revenue"], Decimal("400.00"))
        self.assertEqual(summary["cost_cad"], Decimal("200.00"))
        self.assertEqual(summary["profit"], Decimal("200.00"))
        self.assertEqual(summary["margin_pct"], Decimal("50.00"))

    def test_missing_cost_and_missing_revenue_never_guess_profit(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        missing_cost = self.canada_order(code="POMISSINGCOST")
        self.invoice(
            missing_cost,
            number="INV-PROFIT-NO-COST",
            amount=Decimal("1000"),
            currency="CAD",
            market="north_america",
            region="CA",
        )
        missing_revenue = self.canada_order(
            code="POMISSINGREVENUE",
            cost_bdt=Decimal("8500"),
        )

        rows = {row["production_order_id"]: row for row in self.report()["rows"]}

        self.assertEqual(rows[missing_cost.pk]["data_status"], "Missing cost")
        self.assertIsNone(rows[missing_cost.pk]["profit"])
        self.assertIsNone(rows[missing_cost.pk]["margin_pct"])
        self.assertEqual(rows[missing_revenue.pk]["data_status"], "Missing revenue")
        self.assertIsNone(rows[missing_revenue.pk]["profit"])
        self.assertIsNone(rows[missing_revenue.pk]["margin_pct"])

    def test_missing_exchange_rate_keeps_canada_profit_unavailable(self):
        order = self.canada_order(code="POMISSINGRATE", cost_bdt=Decimal("8500"))
        self.invoice(
            order,
            number="INV-PROFIT-NO-RATE",
            amount=Decimal("200"),
            currency="CAD",
            market="north_america",
            region="CA",
        )

        row = next(
            item for item in self.report()["rows"]
            if item["production_order_id"] == order.pk
        )

        self.assertEqual(row["data_status"], "Missing exchange rate")
        self.assertIsNone(row["cost_cad"])
        self.assertIsNone(row["profit"])
        self.assertIsNone(row["margin_pct"])

    def test_combined_view_converts_local_revenue_and_never_adds_bdt_as_cad(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        canada = self.canada_order(code="POCOMBINEDCA", cost_bdt=Decimal("85000"))
        self.invoice(
            canada,
            number="INV-PROFIT-COMBINED-CA",
            amount=Decimal("2000"),
            currency="CAD",
            market="north_america",
            region="CA",
        )
        local = self.local_order(
            code="POCOMBINEDBD",
            charge=Decimal("850"),
            cost=Decimal("425"),
        )
        self.invoice(
            local,
            number="INV-PROFIT-COMBINED-BD",
            amount=Decimal("85000"),
            currency="BDT",
            market="bangladesh",
            region="BD",
            invoice_type="sewing_charge",
        )

        combined = self.report()["combined"]

        self.assertTrue(combined["complete"])
        self.assertEqual(combined["revenue_cad"], Decimal("3000.00"))
        self.assertEqual(combined["cost_cad"], Decimal("1500.00"))
        self.assertEqual(combined["profit_cad"], Decimal("1500.00"))
        self.assertEqual(combined["margin_pct"], Decimal("50.00"))

    def test_mixed_invoice_currencies_are_unclassified_and_not_combined(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.canada_order(code="POMIXED001", cost_bdt=Decimal("8500"))
        self.invoice(
            order,
            number="INV-PROFIT-MIX-CAD",
            amount=Decimal("100"),
            currency="CAD",
            market="north_america",
            region="CA",
        )
        self.invoice(
            order,
            number="INV-PROFIT-MIX-BDT",
            amount=Decimal("8500"),
            currency="BDT",
            market="bangladesh",
            region="BD",
        )

        report = self.report()
        row = next(item for item in report["rows"] if item["production_order_id"] == order.pk)

        self.assertEqual(row["classification"], "unclassified")
        self.assertEqual(row["data_status"], "Unavailable")
        self.assertIsNone(row["profit"])
        self.assertFalse(report["combined"]["complete"])
        self.assertIsNone(report["combined"]["revenue_cad"])

    def test_friendly_po_renders_and_internal_id_search_finds_order(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.canada_order(code="PO260706120000ABC123", cost_bdt=Decimal("8500"))
        self.invoice(
            order,
            number="INV-PROFIT-SEARCH",
            amount=Decimal("200"),
            currency="CAD",
            market="north_america",
            region="CA",
        )

        response = self.client.get(
            reverse("production_profit_report"),
            {
                "year": self.today.year,
                "month": self.today.month,
                "q": order.internal_order_id,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, order.purchase_order_number)
        self.assertContains(response, "Internal Order ID")
        self.assertEqual(len(response.context["rows"]), 1)
        self.assertEqual(response.context["rows"][0]["production_order_id"], order.pk)

    def test_sample_summary_table_paid_balance_and_friendly_po(self):
        order = self.canada_order(code="PO260706130000SAMPLE", quantity=4)
        sample = self.sample_invoice(
            order,
            number="INV-SAMPLE-SUMMARY",
            amount=Decimal("400"),
            paid_amount=Decimal("125"),
        )

        report = self.report()

        self.assertEqual(report["sample_invoice_count"], 1)
        summary = report["sample_revenue"][0]
        self.assertEqual(summary["currency"], "CAD")
        self.assertEqual(summary["invoice_count"], 1)
        self.assertEqual(summary["piece_count"], 4)
        self.assertEqual(summary["revenue"], Decimal("400.00"))
        self.assertEqual(summary["paid"], Decimal("125.00"))
        self.assertEqual(summary["balance"], Decimal("275.00"))
        row = report["sample_rows"][0]
        self.assertEqual(row["invoice_id"], sample.pk)
        self.assertEqual(row["purchase_order_number"], order.purchase_order_number)
        self.assertEqual(row["payment_status"], "Partially paid")
        self.assertEqual(row["credit_status"], "Credit tracking unavailable")

        response = self.client.get(
            reverse("production_profit_report"),
            {"year": self.today.year, "month": self.today.month},
        )
        self.assertContains(response, "INV-SAMPLE-SUMMARY")
        self.assertContains(response, order.purchase_order_number)
        self.assertContains(
            response,
            "Sample revenue is tracked separately and excluded from production margin.",
        )

    def test_sample_courier_cost_profit_and_margin_use_recorded_invoice_fields(self):
        order = self.canada_order(code="POSAMPLECOST", quantity=3)
        self.sample_invoice(
            order,
            number="INV-SAMPLE-COST",
            amount=Decimal("300"),
            shipping_amount=Decimal("25"),
            sewing_charge=Decimal("70"),
            other_internal_cost=Decimal("30"),
        )

        report = self.report()
        row = report["sample_rows"][0]
        summary = report["sample_revenue"][0]

        self.assertEqual(row["courier_charge"], Decimal("25.00"))
        self.assertEqual(row["sample_cost"], Decimal("100.00"))
        self.assertEqual(row["gross_profit"], Decimal("200.00"))
        self.assertEqual(row["margin_pct"], Decimal("66.67"))
        self.assertEqual(summary["cost"], Decimal("100.00"))
        self.assertEqual(summary["profit"], Decimal("200.00"))
        self.assertEqual(summary["margin_pct"], Decimal("66.67"))

    def test_sample_revenue_is_excluded_from_all_production_profit_totals(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.canada_order(
            code="POSAMPLEEXCLUDED",
            cost_bdt=Decimal("8500"),
        )
        self.invoice(
            order,
            number="INV-BULK-FOR-SAMPLE-EXCLUSION",
            amount=Decimal("200"),
            currency="CAD",
            market="north_america",
            region="CA",
        )
        self.sample_invoice(
            order,
            number="INV-SAMPLE-EXCLUDED",
            amount=Decimal("500"),
        )

        report = self.report()

        self.assertEqual(report["sample_revenue"][0]["revenue"], Decimal("500.00"))
        self.assertEqual(report["canada_export"]["revenue"], Decimal("200.00"))
        self.assertEqual(report["canada_export"]["profit"], Decimal("100.00"))
        self.assertEqual(report["canada_export"]["margin_pct"], Decimal("50.00"))
        self.assertEqual(report["canada_export_sewing"]["order_count"], 0)
        self.assertEqual(report["combined"]["revenue_cad"], Decimal("200.00"))
        self.assertEqual(report["combined"]["margin_pct"], Decimal("50.00"))

    def test_sample_without_existing_quantity_displays_unavailable(self):
        customer = Customer.objects.create(account_brand="Sample Buyer")
        self.sample_invoice(
            number="INV-SAMPLE-NO-QTY",
            customer=customer,
        )

        response = self.client.get(
            reverse("production_profit_report"),
            {"year": self.today.year, "month": self.today.month},
        )

        self.assertEqual(response.status_code, 200)
        row = next(
            item for item in response.context["sample_rows"]
            if item["invoice_number"] == "INV-SAMPLE-NO-QTY"
        )
        self.assertIsNone(row["pieces"])
        self.assertIsNone(response.context["sample_revenue"][0]["piece_count"])
        self.assertContains(response, "Unavailable")

    def test_sample_resolves_lead_opportunity_and_production_links(self):
        customer = Customer.objects.create(account_brand="Linked Sample Buyer")
        lead = Lead.objects.create(
            lead_id="LEAD-SAMPLE-001",
            account_brand="Linked Sample Buyer",
            customer=customer,
        )
        opportunity = Opportunity.objects.create(
            opportunity_id="OPP-SAMPLE-001",
            lead=lead,
            customer=customer,
        )
        order = self.canada_order(
            code="PO260706140000LINKED",
            quantity=2,
            lead=lead,
            opportunity=opportunity,
            customer=customer,
        )
        self.sample_invoice(
            order,
            number="INV-SAMPLE-LINKED",
            customer=customer,
        )

        row = self.report()["sample_rows"][0]

        self.assertEqual(row["lead_id"], lead.lead_id)
        self.assertEqual(row["opportunity_id"], opportunity.opportunity_id)
        self.assertEqual(row["purchase_order_number"], order.purchase_order_number)

    def test_linked_sample_fields_and_product_descriptor_classify_sample(self):
        quick = QuickCosting.objects.create(
            buyer_name="Purpose Sample Buyer",
            project_name="Purpose item",
            product_type="Other",
            costing_purpose=QuickCosting.PURPOSE_SAMPLE,
            quantity=3,
        )
        purpose_invoice = self.invoice(
            number="INV-SAMPLE-PURPOSE",
            amount=Decimal("90"),
            currency="CAD",
            market="north_america",
            region="CA",
            invoice_type="bulk",
            quick_costing=quick,
        )
        descriptor_order = self.canada_order(
            code="POSAMPLEDESCRIPTOR",
            quantity=5,
        )
        descriptor_order.title = "Client sample development"
        descriptor_order.save(update_fields=["title"])
        descriptor_invoice = self.invoice(
            descriptor_order,
            number="INV-SAMPLE-DESCRIPTOR",
            amount=Decimal("125"),
            currency="CAD",
            market="north_america",
            region="CA",
            invoice_type="bulk",
        )

        sample_ids = {row["invoice_id"] for row in self.report()["sample_rows"]}

        self.assertIn(purpose_invoice.pk, sample_ids)
        self.assertIn(descriptor_invoice.pk, sample_ids)

    def test_linked_production_sample_type_classifies_sample(self):
        order = self.canada_order(
            code="PO260706150000SAMPLING",
            quantity=6,
            production_order_type="sampling",
        )
        invoice = self.invoice(
            order,
            number="INV-SAMPLE-PRODUCTION-TYPE",
            amount=Decimal("180"),
            currency="CAD",
            market="north_america",
            region="CA",
            invoice_type="bulk",
        )

        report = self.report()
        row = next(item for item in report["sample_rows"] if item["invoice_id"] == invoice.pk)

        self.assertIn("Production sample", row["sample_type"])
        self.assertNotIn(order.pk, {item["production_order_id"] for item in report["rows"]})

    def test_conflicting_sample_and_sewing_signal_is_unclassified(self):
        order = self.canada_order(
            code="POCONFLICTINGSAMPLE",
            production_order_type="sampling",
        )
        invoice = self.invoice(
            order,
            number="INV-CONFLICTING-SAMPLE",
            amount=Decimal("100"),
            currency="CAD",
            market="north_america",
            region="CA",
            invoice_type="sewing_charge",
        )

        report = self.report()

        self.assertEqual(report["sample_rows"], [])
        self.assertEqual(report["unclassified_sample_invoices"], [{
            "invoice_id": invoice.pk,
            "invoice_number": invoice.invoice_number,
            "reason": "Conflicting sample and sewing-charge classification",
        }])
        self.assertNotIn(order.pk, {item["production_order_id"] for item in report["rows"]})

    def test_bulk_and_sewing_charge_invoices_do_not_enter_sample_section(self):
        bulk = self.canada_order(code="POBULKNOTSAMPLE")
        sewing = self.canada_order(code="POSEWINGNOTSAMPLE")
        bulk_invoice = self.invoice(
            bulk,
            number="INV-BULK-NOT-SAMPLE",
            amount=Decimal("200"),
            currency="CAD",
            market="north_america",
            region="CA",
        )
        sewing_invoice = self.invoice(
            sewing,
            number="INV-SEWING-NOT-SAMPLE",
            amount=Decimal("100"),
            currency="CAD",
            market="north_america",
            region="CA",
            invoice_type="sewing_charge",
        )

        sample_ids = {row["invoice_id"] for row in self.report()["sample_rows"]}

        self.assertNotIn(bulk_invoice.pk, sample_ids)
        self.assertNotIn(sewing_invoice.pk, sample_ids)

    def test_year_and_month_filters_apply_to_sample_invoices(self):
        previous_year = self.today.year - 1
        previous_date = self.today.replace(year=previous_year)
        current = self.sample_invoice(number="INV-SAMPLE-CURRENT")
        previous = self.sample_invoice(
            number="INV-SAMPLE-PREVIOUS",
            issue_date=previous_date,
        )

        current_ids = {row["invoice_id"] for row in self.report()["sample_rows"]}
        previous_report = build_production_profit_report(
            year=previous_year,
            month=previous_date.month,
        )
        previous_ids = {row["invoice_id"] for row in previous_report["sample_rows"]}

        self.assertIn(current.pk, current_ids)
        self.assertNotIn(previous.pk, current_ids)
        self.assertEqual(previous_ids, {previous.pk})

    def test_sample_only_order_is_not_added_to_production_rows(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.canada_order(
            code="POSAMPLEONLY",
            cost_bdt=Decimal("8500"),
        )
        self.sample_invoice(order, number="INV-SAMPLE-ONLY")

        report = self.report()

        self.assertEqual(report["sample_invoice_count"], 1)
        self.assertNotIn(order.pk, {row["production_order_id"] for row in report["rows"]})
        self.assertEqual(report["canada_export"]["order_count"], 0)
        self.assertFalse(report["combined"]["complete"])

    def test_bulk_sewing_sample_and_other_revenue_are_independent(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        customer = Customer.objects.create(
            account_brand="Revenue Brand",
            contact_name="Revenue Client",
            country="Canada",
        )
        bulk = self.canada_order(
            code="POREVENUEBULK",
            cost_bdt=Decimal("8500"),
            customer=customer,
        )
        self.invoice(
            bulk,
            number="INV-REVENUE-BULK",
            amount=Decimal("200"),
            currency="CAD",
            market="north_america",
            region="CA",
            customer=customer,
        )
        sewing = self.canada_order(
            code="POREVENUESEWING",
            cost_bdt=Decimal("8500"),
            sewing_cost_bdt=Decimal("4250"),
            customer=customer,
        )
        self.invoice(
            sewing,
            number="INV-REVENUE-SEWING",
            amount=Decimal("100"),
            currency="CAD",
            market="north_america",
            region="CA",
            invoice_type="sewing_charge",
            customer=customer,
        )
        self.sample_invoice(
            number="INV-REVENUE-SAMPLE",
            amount=Decimal("50"),
            customer=customer,
            sewing_charge=Decimal("20"),
        )
        self.invoice(
            number="INV-REVENUE-OTHER",
            amount=Decimal("75"),
            currency="CAD",
            market="north_america",
            region="CA",
            customer=customer,
            notes="Tech Pack service fee",
        )

        report = self.report()

        self.assertEqual(report["bulk_revenue"][0]["revenue"], Decimal("200.00"))
        self.assertEqual(report["sewing_revenue"][0]["revenue"], Decimal("100.00"))
        self.assertEqual(report["sample_revenue"][0]["revenue"], Decimal("50.00"))
        self.assertEqual(report["other_revenue"][0]["revenue"], Decimal("75.00"))
        self.assertEqual(report["canada_export"]["revenue"], Decimal("300.00"))
        self.assertEqual(report["combined"]["revenue_cad"], Decimal("300.00"))

    def test_accounting_other_revenue_and_reconciliation_are_explicit(self):
        customer = Customer.objects.create(
            account_brand="Service Brand",
            contact_name="Service Client",
            country="Canada",
        )
        self.invoice(
            number="INV-TECH-PACK",
            amount=Decimal("500"),
            currency="CAD",
            market="north_america",
            region="CA",
            paid_amount=Decimal("100"),
            customer=customer,
            notes="Tech Pack service fee",
        )
        AccountingEntry.objects.create(
            date=self.today,
            side="CA",
            direction="IN",
            main_type="INCOME",
            sub_type="Design fee",
            currency="CAD",
            amount_original=Decimal("250"),
            customer=customer,
        )

        report = self.report()

        self.assertEqual(report["other_revenue"][0]["revenue"], Decimal("750.00"))
        self.assertEqual({row["service_type"] for row in report["other_rows"]}, {"Tech Pack", "Design"})
        self.assertFalse(report["company_revenue"][0]["cost_available"])
        reconciliation = report["accounting_reconciliation"][0]
        self.assertEqual(reconciliation["categorized_revenue"], Decimal("750.00"))
        self.assertEqual(reconciliation["accounting_revenue"], Decimal("250.00"))
        self.assertEqual(reconciliation["difference"], Decimal("500.00"))

    def test_categorized_revenue_reconciles_to_matching_accounting_income(self):
        self.sample_invoice(
            number="INV-RECONCILED-SAMPLE",
            amount=Decimal("100"),
        )
        AccountingEntry.objects.create(
            date=self.today,
            side="CA",
            direction="IN",
            main_type="INCOME",
            sub_type="Invoice payment received",
            currency="CAD",
            amount_original=Decimal("100"),
        )

        reconciliation = self.report()["accounting_reconciliation"][0]

        self.assertEqual(reconciliation["categorized_revenue"], Decimal("100.00"))
        self.assertEqual(reconciliation["accounting_revenue"], Decimal("100.00"))
        self.assertEqual(reconciliation["difference"], Decimal("0.00"))

    def test_client_brand_country_date_and_revenue_type_filters_apply(self):
        alpha = Customer.objects.create(
            account_brand="Alpha Brand",
            contact_name="Alpha Client",
            country="Canada",
        )
        beta = Customer.objects.create(
            account_brand="Beta Brand",
            contact_name="Beta Client",
            country="USA",
        )
        self.sample_invoice(
            number="INV-FILTER-ALPHA",
            customer=alpha,
            issue_date=self.today,
        )
        self.sample_invoice(
            number="INV-FILTER-BETA",
            customer=beta,
            issue_date=self.today,
        )

        report = build_production_profit_report(
            year=self.today.year,
            month=self.today.month,
            start_date=self.today,
            end_date=self.today,
            client="Alpha",
            brand="Alpha Brand",
            country="Canada",
            revenue_type="sample",
        )

        self.assertEqual([row["invoice_number"] for row in report["sample_rows"]], ["INV-FILTER-ALPHA"])
        self.assertTrue(report["export_rows"])
        self.assertEqual({row["revenue_type"] for row in report["export_rows"]}, {"Sample"})

    def test_pdf_and_excel_exports_use_filtered_read_only_report(self):
        self.sample_invoice(number="INV-EXPORT-SAMPLE")

        excel = self.client.get(
            reverse("production_profit_report"),
            {"year": self.today.year, "month": self.today.month, "export": "xlsx"},
        )
        pdf = self.client.get(
            reverse("production_profit_report"),
            {"year": self.today.year, "month": self.today.month, "export": "pdf"},
        )

        self.assertEqual(excel.status_code, 200)
        self.assertEqual(
            excel["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        from openpyxl import load_workbook
        from io import BytesIO
        workbook = load_workbook(BytesIO(excel.content), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Revenue Summary", "Revenue Detail"])
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))

    def test_render_is_read_only_and_does_not_create_exchange_rate(self):
        order = self.canada_order(code="POREADONLY", cost_bdt=Decimal("8500"))
        self.sample_invoice(order, number="INV-SAMPLE-READ-ONLY")
        models = (ProductionOrder, Invoice, AccountingEntry, QuickCosting, ExchangeRate)
        before = {model: model.objects.count() for model in models}

        with CaptureQueriesContext(connection) as captured:
            response = self.client.get(
                reverse("production_profit_report"),
                {"year": self.today.year, "month": self.today.month},
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Exchange rate unavailable")
        self.assertEqual({model: model.objects.count() for model in models}, before)
        write_sql = [
            query["sql"]
            for query in captured
            if query["sql"].lstrip().upper().startswith(("INSERT", "UPDATE", "DELETE"))
        ]
        self.assertEqual(write_sql, [])

    def test_report_query_count_is_bounded_and_has_no_n_plus_one(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        first = self.canada_order(code="POQUERY001", cost_bdt=Decimal("8500"))
        self.invoice(
            first,
            number="INV-PROFIT-QUERY-1",
            amount=Decimal("200"),
            currency="CAD",
            market="north_america",
            region="CA",
        )
        with CaptureQueriesContext(connection) as one_order_queries:
            self.report()

        for index in range(2, 7):
            order = self.canada_order(
                code=f"POQUERY{index:03d}",
                cost_bdt=Decimal("8500"),
            )
            self.invoice(
                order,
                number=f"INV-PROFIT-QUERY-{index}",
                amount=Decimal("200"),
                currency="CAD",
                market="north_america",
                region="CA",
            )

        with CaptureQueriesContext(connection) as six_order_queries:
            self.report()

        self.assertLessEqual(len(one_order_queries), 5)
        self.assertEqual(len(six_order_queries), len(one_order_queries))

    def test_sample_rows_do_not_add_n_plus_one_queries(self):
        ExchangeRate.objects.create(cad_to_bdt=Decimal("85"))
        order = self.canada_order(code="POSAMPLEQUERY", cost_bdt=Decimal("8500"))
        self.invoice(
            order,
            number="INV-SAMPLE-QUERY-BULK",
            amount=Decimal("200"),
            currency="CAD",
            market="north_america",
            region="CA",
        )
        self.sample_invoice(order, number="INV-SAMPLE-QUERY-1")
        with CaptureQueriesContext(connection) as one_sample_queries:
            self.report()

        for index in range(2, 7):
            self.sample_invoice(order, number=f"INV-SAMPLE-QUERY-{index}")

        with CaptureQueriesContext(connection) as six_sample_queries:
            self.report()

        self.assertLessEqual(len(one_sample_queries), 5)
        self.assertEqual(len(six_sample_queries), len(one_sample_queries))
