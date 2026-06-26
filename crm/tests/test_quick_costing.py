import re
from io import BytesIO
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from crm.forms_costing import QuickCostingForm
from crm.models import CostingHeader, Invoice, Lead, Opportunity, QuickCosting


class QuickCostingTests(TestCase):
    def _admin_user(self, username="quick-costing-admin"):
        user_model = get_user_model()
        return user_model.objects.create_superuser(
            username=username,
            email=f"{username}@example.com",
            password="test-pass",
        )

    def _opportunity(self):
        lead = Lead.objects.create(
            account_brand="Test Streetwear Co",
            contact_name="Taylor Buyer",
            email="buyer@example.com",
            product_category="Hoodie",
            primary_product_type="Streetwear",
            order_quantity="300",
        )
        return Opportunity.objects.create(
            lead=lead,
            product_category="Hoodie",
            product_type="Streetwear",
            moq_units=300,
        )

    def _quick_costing(self, **overrides):
        data = {
            "buyer_name": "Test Buyer",
            "project_name": "Fast Hoodie",
            "product_type": "Streetwear",
            "quantity": 100,
            "currency": None,
            "exchange_rate_bdt_per_cad": Decimal("90.00"),
            "material_cost": Decimal("500.00"),
            "production_cost": Decimal("300.00"),
            "other_expenses": Decimal("200.00"),
            "shipping_cost": Decimal("100.00"),
            "selling_price_per_piece": Decimal("15.00"),
            "commission_per_piece": Decimal("1.00"),
            "target_margin_percent": Decimal("20.00"),
        }
        data.update(overrides)
        return QuickCosting.objects.create(**data)

    def _costing_user(self, username="quick-costing-staff", approve=False):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username=username,
            email=f"{username}@example.com",
            password="test-pass",
        )
        access = user.access
        access.can_costing = True
        access.can_view_internal_costing = True
        access.can_costing_approve = approve
        access.save()
        return user

    def test_calculation_summary(self):
        quick = QuickCosting(
            buyer_name="Test Buyer",
            project_name="Fast Hoodie",
            product_type="Streetwear",
            quantity=100,
            exchange_rate_bdt_per_cad=Decimal("90.00"),
            material_cost=Decimal("500.00"),
            production_cost=Decimal("300.00"),
            other_expenses=Decimal("200.00"),
            shipping_cost=Decimal("100.00"),
            selling_price_per_piece=Decimal("15.00"),
            commission_per_piece=Decimal("1.00"),
            target_margin_percent=Decimal("20.00"),
        )

        summary = quick.calculation_summary()

        self.assertEqual(summary["total_cost"], Decimal("1100.00"))
        self.assertEqual(summary["cost_per_piece"], Decimal("11.00"))
        self.assertEqual(summary["revenue"], Decimal("1500.00"))
        self.assertEqual(summary["profit_per_piece"], Decimal("4.00"))
        self.assertEqual(summary["total_profit"], Decimal("400.00"))
        self.assertEqual(summary["profit_margin_percent"], Decimal("26.66666666666666666666666667"))
        self.assertEqual(summary["commission_total"], Decimal("100.00"))
        self.assertEqual(summary["net_profit_per_piece"], Decimal("3.00"))
        self.assertEqual(summary["net_profit_total"], Decimal("300.00"))
        self.assertEqual(summary["net_profit_margin_percent"], Decimal("20.0"))
        self.assertEqual(summary["margin_status"], "Meets target")

    def test_detailed_per_piece_cost_and_percentage_commission(self):
        quick = QuickCosting(
            buyer_name="Detailed Buyer",
            project_name="Detailed Hoodie",
            product_type="Streetwear",
            quantity=100,
            currency="CAD",
            fabric_cost_per_kg=Decimal("100.00"),
            fabric_consumption_kg_per_piece=Decimal("0.5000"),
            making_cost_per_piece=Decimal("20.00"),
            print_embroidery_cost_per_piece=Decimal("5.00"),
            trims_cost_per_piece=Decimal("3.00"),
            packaging_cost_per_piece=Decimal("2.00"),
            other_expenses=Decimal("100.00"),
            shipping_cost=Decimal("200.00"),
            selling_price_per_piece=Decimal("100.00"),
            commission_percent=Decimal("5.00"),
        )

        summary = quick.calculation_summary()

        self.assertTrue(summary["uses_detailed_costing"])
        self.assertEqual(summary["fabric_cost_per_piece"], Decimal("50.000000"))
        self.assertEqual(summary["other_expenses_per_piece"], Decimal("1.00"))
        self.assertEqual(summary["shipping_cost_per_piece"], Decimal("2.00"))
        self.assertEqual(summary["cost_per_piece"], Decimal("83.000000"))
        self.assertEqual(summary["total_cost"], Decimal("8300.000000"))
        self.assertEqual(summary["commission_per_piece"], Decimal("5.00"))
        self.assertEqual(summary["commission_total"], Decimal("500.00"))
        self.assertEqual(summary["gross_profit_per_piece"], Decimal("17.000000"))
        self.assertEqual(summary["net_profit_per_piece"], Decimal("12.000000"))
        self.assertEqual(summary["net_profit_total"], Decimal("1200.000000"))
        self.assertEqual(summary["net_profit_margin_percent"], Decimal("12.000000"))

    def test_percentage_commission_is_not_double_counted(self):
        quick = QuickCosting(
            buyer_name="Commission Buyer",
            project_name="Commission Check",
            product_type="Streetwear",
            quantity=10,
            currency="USD",
            fabric_cost_per_kg=Decimal("100.00"),
            fabric_consumption_kg_per_piece=Decimal("0.5000"),
            making_cost_per_piece=Decimal("30.00"),
            selling_price_per_piece=Decimal("100.00"),
            commission_percent=Decimal("5.00"),
            commission_per_piece=Decimal("25.00"),
        )

        summary = quick.calculation_summary()

        self.assertEqual(summary["cost_per_piece"], Decimal("80.000000"))
        self.assertEqual(summary["gross_profit_per_piece"], Decimal("20.000000"))
        self.assertEqual(summary["commission_per_piece"], Decimal("5.00"))
        self.assertEqual(summary["commission_total"], Decimal("50.00"))
        self.assertEqual(summary["net_profit_per_piece"], Decimal("15.000000"))
        self.assertEqual(summary["net_profit_total"], Decimal("150.000000"))

    def test_legacy_absolute_commission_still_works(self):
        quick = self._quick_costing(commission_percent=None, commission_per_piece=Decimal("1.25"))

        summary = quick.calculation_summary()

        self.assertFalse(summary["uses_detailed_costing"])
        self.assertTrue(summary["is_legacy_currency"])
        self.assertEqual(summary["commission_per_piece"], Decimal("1.25"))
        self.assertEqual(summary["commission_total"], Decimal("125.00"))
        self.assertEqual(summary["net_profit_per_piece"], Decimal("2.75"))

    def test_bdt_cad_and_usd_currency_labels(self):
        admin = self._admin_user("quick-costing-currency-admin")
        self.client.force_login(admin)

        expected_labels = {
            "BDT": "৳100.00 BDT",
            "CAD": "CAD $100.00",
            "USD": "USD $100.00",
        }
        for currency, expected_label in expected_labels.items():
            with self.subTest(currency=currency):
                quick = QuickCosting.objects.create(
                    buyer_name=f"{currency} Buyer",
                    project_name=f"{currency} Costing",
                    product_type="Streetwear",
                    quantity=10,
                    currency=currency,
                    fabric_cost_per_kg=Decimal("100.00"),
                    fabric_consumption_kg_per_piece=Decimal("0.5000"),
                    making_cost_per_piece=Decimal("10.00"),
                    selling_price_per_piece=Decimal("100.00"),
                    commission_percent=Decimal("5.00"),
                )

                response = self.client.get(reverse("quick_costing_detail", args=[quick.pk]))

                self.assertEqual(response.status_code, 200)
                self.assertContains(response, expected_label)
                self.assertContains(response, f"<strong>{currency}</strong>", html=True)
                self.assertContains(response, "5.00%")

    def test_new_form_has_explicit_cost_basis_and_currency_labels(self):
        form = QuickCostingForm()

        self.assertIn("currency", form.fields)
        self.assertEqual(form.fields["fabric_cost_per_kg"].label, "Fabric Cost Per KG")
        self.assertIn("per kg", form.fields["fabric_cost_per_kg"].help_text)
        self.assertIn("per piece", form.fields["making_cost_per_piece"].help_text)
        self.assertIn("total order", form.fields["other_expenses"].help_text)
        self.assertIn("Percentage", form.fields["commission_percent"].help_text)
        self.assertTrue(form.fields["commission_per_piece"].disabled)
        self.assertEqual(form.fields["commission_per_piece"].widget.input_type, "hidden")

    def test_form_requires_fabric_cost_and_consumption_together(self):
        form = QuickCostingForm(
            data={
                "buyer_name": "Fabric Buyer",
                "project_name": "Fabric Check",
                "product_type": "Streetwear",
                "costing_purpose": "bulk",
                "quantity": "100",
                "currency": "BDT",
                "fabric_cost_per_kg": "100.00",
                "fabric_consumption_kg_per_piece": "",
                "making_cost_per_piece": "10.00",
                "print_embroidery_cost_per_piece": "0.00",
                "trims_cost_per_piece": "0.00",
                "packaging_cost_per_piece": "0.00",
                "other_expenses": "0.00",
                "shipping_cost": "0.00",
                "selling_price_per_piece": "100.00",
                "commission_percent": "5.00",
                "target_margin_percent": "20.00",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("fabric_consumption_kg_per_piece", form.errors)

    def test_calculation_summary_handles_missing_exchange_and_zero_quantity(self):
        quick = QuickCosting(
            buyer_name="Test Buyer",
            project_name="Zero Quantity Safety",
            product_type="Streetwear",
            quantity=0,
            material_cost=Decimal("500.00"),
            production_cost=Decimal("300.00"),
            other_expenses=Decimal("200.00"),
            shipping_cost=Decimal("100.00"),
            selling_price_per_piece=Decimal("15.00"),
        )

        summary = quick.calculation_summary()

        self.assertIsNone(summary["exchange_rate"])
        self.assertEqual(summary["cost_per_piece"], Decimal("0"))
        self.assertEqual(summary["material_cost_per_piece"], Decimal("0"))
        self.assertEqual(summary["gross_profit_margin_percent"], Decimal("0"))
        self.assertEqual(summary["net_profit_margin_percent"], Decimal("0"))

    def test_form_blocks_zero_quantity_and_negative_cost(self):
        form = QuickCostingForm(
            data={
                "buyer_name": "Test Buyer",
                "project_name": "Fast Hoodie",
                "product_type": "Streetwear",
                "quantity": 0,
                "currency": "BDT",
                "exchange_rate_bdt_per_cad": "0",
                "fabric_cost_per_kg": "-1.00",
                "fabric_consumption_kg_per_piece": "0.5000",
                "making_cost_per_piece": "0.00",
                "print_embroidery_cost_per_piece": "0.00",
                "trims_cost_per_piece": "0.00",
                "packaging_cost_per_piece": "0.00",
                "other_expenses": "0.00",
                "shipping_cost": "",
                "selling_price_per_piece": "15.00",
                "commission_percent": "101.00",
                "target_margin_percent": "-1",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("quantity", form.errors)
        self.assertIn("fabric_cost_per_kg", form.errors)
        self.assertIn("exchange_rate_bdt_per_cad", form.errors)
        self.assertIn("commission_percent", form.errors)
        self.assertIn("target_margin_percent", form.errors)

    def test_detail_handles_missing_exchange_rate(self):
        admin = self._admin_user("quick-costing-no-rate-admin")
        self.client.force_login(admin)
        quick = QuickCosting.objects.create(
            buyer_name="Old Buyer",
            project_name="Legacy Quick Costing",
            product_type="Streetwear",
            quantity=100,
            currency=None,
            material_cost=Decimal("500.00"),
            production_cost=Decimal("300.00"),
            other_expenses=Decimal("200.00"),
            shipping_cost=Decimal("100.00"),
            selling_price_per_piece=Decimal("15.00"),
        )

        response = self.client.get(reverse("quick_costing_detail", args=[quick.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Exchange Rate")
        self.assertContains(response, "N/A")
        self.assertContains(response, "Bulk Production Costing")
        self.assertContains(response, "৳1,500.00 BDT")
        self.assertContains(response, "CAD N/A")

    def test_quick_costing_create_detail_and_list(self):
        admin = self._admin_user()
        self.client.force_login(admin)

        create_response = self.client.post(
            reverse("cost_sheet_create"),
            data={
                "costing_type": "quick",
                "buyer_name": "Test Buyer",
                "project_name": "Fast Hoodie",
                "product_type": "Streetwear",
                "costing_purpose": "bulk",
                "quantity": 100,
                "currency": "BDT",
                "exchange_rate_bdt_per_cad": "90.00",
                "fabric_cost_per_kg": "10.00",
                "fabric_consumption_kg_per_piece": "0.5000",
                "making_cost_per_piece": "2.00",
                "print_embroidery_cost_per_piece": "0.50",
                "trims_cost_per_piece": "0.30",
                "packaging_cost_per_piece": "0.20",
                "other_expenses": "200.00",
                "shipping_cost": "100.00",
                "selling_price_per_piece": "20.00",
                "commission_percent": "5.00",
                "target_margin_percent": "20.00",
            },
        )

        quick = QuickCosting.objects.get(project_name="Fast Hoodie")
        self.assertEqual(create_response.status_code, 302)
        self.assertEqual(create_response["Location"], reverse("quick_costing_detail", args=[quick.pk]))
        self.assertEqual(quick.costing_type, "quick")
        self.assertEqual(quick.costing_purpose, QuickCosting.PURPOSE_BULK)

        detail_response = self.client.get(reverse("quick_costing_detail", args=[quick.pk]))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "Quick Costing")
        self.assertContains(detail_response, "Costing Purpose")
        self.assertContains(detail_response, "Bulk Production Costing")
        self.assertContains(detail_response, "Approval Status")
        self.assertContains(detail_response, "Pending")
        self.assertContains(detail_response, reverse("quick_costing_export_pdf", args=[quick.pk]))
        self.assertContains(detail_response, reverse("quick_costing_export_excel", args=[quick.pk]))
        self.assertContains(detail_response, reverse("quick_costing_edit", args=[quick.pk]))
        self.assertContains(detail_response, "Draft")
        self.assertContains(detail_response, "Approve costing before creating quotation.")
        self.assertContains(detail_response, "Shipping Cost")
        self.assertContains(detail_response, "Exchange Rate")
        self.assertContains(detail_response, "1 CAD = 90.00 BDT")
        self.assertContains(detail_response, "Profit Before Commission")
        self.assertContains(detail_response, "Final Profit After Commission")
        self.assertContains(detail_response, "Commission")
        self.assertContains(detail_response, "Meets target")
        self.assertContains(detail_response, "৳2,000.00 BDT")
        self.assertContains(detail_response, "৳1,100.00 BDT")
        self.assertContains(detail_response, "৳800.00 BDT")

        pdf_response = self.client.get(reverse("quick_costing_export_pdf", args=[quick.pk]))
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")
        pdf_content = pdf_response.content

        def assert_pdf_contains(value):
            self.assertTrue(value in pdf_content, f"PDF did not contain {value!r}")

        assert_pdf_contains(b"COSTING SHEET")
        assert_pdf_contains(b"BUYER NAME")
        assert_pdf_contains(b"PROJECT NAME")
        assert_pdf_contains(b"PRODUCT TYPE")
        assert_pdf_contains(b"QUANTITY")
        assert_pdf_contains(b"EXCHANGE RATE")
        assert_pdf_contains(b"Per Piece - BDT")
        assert_pdf_contains(b"Total Order - BDT")
        assert_pdf_contains(b"Fabric Cost")
        assert_pdf_contains(b"Making and Finishing")
        assert_pdf_contains(b"Other Expenses")
        assert_pdf_contains(b"Shipping Cost")
        assert_pdf_contains(b"Total Cost")
        assert_pdf_contains(b"COST PER PIECE")
        assert_pdf_contains(b"SELLING PRICE PER PIECE")
        assert_pdf_contains(b"TOTAL ORDER VALUE")
        assert_pdf_contains(b"Profit Before Commission")
        assert_pdf_contains(b"COMMISSION PER PIECE")
        assert_pdf_contains(b"COMMISSION TOTAL")
        assert_pdf_contains(b"Final Profit After Commission")
        assert_pdf_contains(b"GROSS PROFIT MARGIN")
        assert_pdf_contains(b"NET PROFIT MARGIN")
        assert_pdf_contains(b"TARGET MARGIN")
        assert_pdf_contains(b"MARGIN STATUS")
        assert_pdf_contains(b"Meets target")
        assert_pdf_contains(b"PREPARED BY")
        assert_pdf_contains(b"Thank You!")
        assert_pdf_contains(b"For Your Business")
        assert_pdf_contains(b"100.00")
        assert_pdf_contains(b"1,100.00")
        assert_pdf_contains(b"800.00")
        self.assertFalse(re.search(rb"0\.9254\d*\s+0\.2823\d*\s+0\.6", pdf_content))

        list_response = self.client.get(reverse("cost_sheet_list") + "?costing_type=quick")
        self.assertEqual(list_response.status_code, 200)
        self.assertContains(list_response, "Quick")
        self.assertContains(list_response, "Draft")
        self.assertContains(list_response, "Edit")
        self.assertContains(list_response, "Excel")
        self.assertContains(list_response, "PDF")
        self.assertContains(list_response, "Fast Hoodie")
        self.assertContains(list_response, "Bulk Production")
        self.assertContains(list_response, "BDT")
        self.assertContains(list_response, "৳1,100.00 BDT")

    def test_quick_costing_can_be_created_from_opportunity(self):
        admin = self._admin_user("quick-costing-opportunity-admin")
        opportunity = self._opportunity()
        self.client.force_login(admin)

        response = self.client.post(
            reverse("cost_sheet_create_for_opportunity", args=[opportunity.pk]),
            data={
                "costing_type": "quick",
                "buyer_name": "Test Streetwear Co",
                "project_name": "Oversized Hoodie",
                "product_type": "Streetwear",
                "costing_purpose": "sample",
                "quantity": 300,
                "currency": "BDT",
                "exchange_rate_bdt_per_cad": "90.00",
                "fabric_cost_per_kg": "100.00",
                "fabric_consumption_kg_per_piece": "0.5000",
                "making_cost_per_piece": "50.00",
                "print_embroidery_cost_per_piece": "10.00",
                "trims_cost_per_piece": "5.00",
                "packaging_cost_per_piece": "5.00",
                "other_expenses": "2000.00",
                "shipping_cost": "5000.00",
                "selling_price_per_piece": "600.00",
                "commission_percent": "5.00",
                "target_margin_percent": "20.00",
            },
        )

        quick = QuickCosting.objects.get(project_name="Oversized Hoodie")
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("quick_costing_detail", args=[quick.pk]))
        self.assertEqual(quick.opportunity, opportunity)
        self.assertEqual(quick.account_brand, "Test Streetwear Co")
        self.assertEqual(quick.contact_name, "Taylor Buyer")
        self.assertEqual(quick.costing_purpose, QuickCosting.PURPOSE_SAMPLE)

        detail_response = self.client.get(reverse("quick_costing_detail", args=[quick.pk]))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, reverse("opportunity_detail", args=[opportunity.pk]))
        self.assertContains(detail_response, opportunity.opportunity_id)
        self.assertContains(detail_response, "Account / Brand")
        self.assertContains(detail_response, "Test Streetwear Co")
        self.assertContains(detail_response, "Taylor Buyer")
        self.assertContains(detail_response, "Sample Costing")

    def test_opportunity_detail_lists_quick_costings_and_status(self):
        admin = self._admin_user("quick-costing-opportunity-list-admin")
        opportunity = self._opportunity()
        self.client.force_login(admin)
        quick = QuickCosting.objects.create(
            opportunity=opportunity,
            account_brand="Test Streetwear Co",
            contact_name="Taylor Buyer",
            buyer_name="Test Streetwear Co",
            project_name="Oversized Hoodie",
            product_type="Streetwear",
            costing_purpose=QuickCosting.PURPOSE_SAMPLE,
            quantity=300,
            currency=None,
            exchange_rate_bdt_per_cad=Decimal("90.00"),
            material_cost=Decimal("25000.00"),
            production_cost=Decimal("15000.00"),
            other_expenses=Decimal("2000.00"),
            shipping_cost=Decimal("5000.00"),
            selling_price_per_piece=Decimal("600.00"),
            commission_per_piece=Decimal("30.00"),
            target_margin_percent=Decimal("20.00"),
            created_by=admin,
        )

        response = self.client.get(reverse("opportunity_detail", args=[opportunity.pk]))
        html = response.content.decode("utf-8")
        timeline_html = html.split("Workflow Activity Timeline", 1)[1].split("</section>", 1)[0]

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Quick Costings")
        self.assertContains(response, "Latest Costing")
        self.assertContains(response, "Quick Costing")
        self.assertContains(response, "Sample Costing")
        self.assertContains(response, "Draft")
        self.assertContains(response, "68.89%")
        self.assertContains(response, "20.00%")
        self.assertContains(response, "Meets target")
        self.assertContains(response, f"QC-{quick.pk}")
        self.assertContains(response, reverse("quick_costing_detail", args=[quick.pk]))
        self.assertContains(response, "৳47,000.00 BDT")
        self.assertContains(response, "CAD $522.22")
        self.assertContains(response, "৳180,000.00 BDT")
        self.assertContains(response, "CAD $2,000.00")
        self.assertContains(response, "৳124,000.00 BDT")
        self.assertContains(response, "CAD $1,377.78")
        self.assertContains(response, "68.89%")
        self.assertIn(f"QC-{quick.pk}", timeline_html)
        self.assertIn("Quick Costing", timeline_html)
        self.assertIn(reverse("quick_costing_detail", args=[quick.pk]), timeline_html)

    def test_opportunity_order_summary_uses_preferred_quick_costing_display_only(self):
        admin = self._admin_user("quick-costing-opportunity-summary-admin")
        opportunity = self._opportunity()
        opportunity.order_value = Decimal("999999.00")
        opportunity.order_value_usd = Decimal("9999.00")
        opportunity.save(update_fields=["order_value", "order_value_usd"])
        self.client.force_login(admin)
        draft = self._quick_costing(
            opportunity=opportunity,
            project_name="Newer Draft Bulk",
            costing_purpose=QuickCosting.PURPOSE_BULK,
            selling_price_per_piece=Decimal("25.00"),
            status=QuickCosting.STATUS_DRAFT,
            created_by=admin,
        )
        approved = self._quick_costing(
            opportunity=opportunity,
            project_name="Approved Sample",
            costing_purpose=QuickCosting.PURPOSE_SAMPLE,
            selling_price_per_piece=Decimal("15.00"),
            status=QuickCosting.STATUS_APPROVED,
            created_by=admin,
        )
        draft.updated_at = timezone.now()
        draft.save(update_fields=["updated_at"])

        response = self.client.get(reverse("opportunity_detail", args=[opportunity.pk]))
        html = response.content.decode("utf-8")
        summary_html = html.split("Order Summary", 1)[1].split("</section>", 1)[0]

        self.assertEqual(response.status_code, 200)
        self.assertIn(f"QC-{approved.pk}", html)
        self.assertIn("৳1,500.00 BDT", summary_html)
        self.assertIn("CAD $16.67", summary_html)
        self.assertIn("Sample Costing", summary_html)
        self.assertNotIn("৳999,999.00 BDT", summary_html)
        opportunity.refresh_from_db()
        self.assertEqual(opportunity.order_value, Decimal("999999.00"))

    def test_opportunity_timeline_uses_latest_costing_when_advanced_and_quick_exist(self):
        admin = self._admin_user("quick-costing-multiple-timeline-admin")
        opportunity = self._opportunity()
        self.client.force_login(admin)
        CostingHeader.objects.create(
            opportunity=opportunity,
            buyer="Test Streetwear Co",
            brand="Test Streetwear Co",
            product_type="Streetwear",
            order_quantity=300,
            moq=300,
        )
        quick = QuickCosting.objects.create(
            opportunity=opportunity,
            account_brand="Test Streetwear Co",
            contact_name="Taylor Buyer",
            buyer_name="Test Streetwear Co",
            project_name="Latest Hoodie Quick",
            product_type="Streetwear",
            quantity=300,
            material_cost=Decimal("25000.00"),
            production_cost=Decimal("15000.00"),
            other_expenses=Decimal("2000.00"),
            shipping_cost=Decimal("5000.00"),
            selling_price_per_piece=Decimal("600.00"),
            created_by=admin,
        )

        response = self.client.get(reverse("opportunity_detail", args=[opportunity.pk]))
        html = response.content.decode("utf-8")
        timeline_html = html.split("Workflow Activity Timeline", 1)[1].split("</section>", 1)[0]

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Multiple Costings")
        self.assertContains(response, "1 advanced · 1 quick")
        self.assertIn(f"QC-{quick.pk}", timeline_html)
        self.assertIn("Quick Costing · 2 total costings", timeline_html)
        self.assertIn(reverse("quick_costing_detail", args=[quick.pk]), timeline_html)

    def test_quick_costing_edit_recalculates(self):
        admin = self._admin_user("quick-costing-edit-admin")
        quick = self._quick_costing()
        self.client.force_login(admin)

        response = self.client.post(
            reverse("quick_costing_edit", args=[quick.pk]),
            data={
                "buyer_name": "Updated Buyer",
                "project_name": "Updated Hoodie",
                "product_type": "Streetwear",
                "costing_purpose": "sample",
                "quantity": 200,
                "exchange_rate_bdt_per_cad": "100.00",
                "material_cost": "1000.00",
                "production_cost": "500.00",
                "other_expenses": "200.00",
                "shipping_cost": "300.00",
                "selling_price_per_piece": "20.00",
                "commission_per_piece": "2.00",
                "target_margin_percent": "25.00",
            },
        )

        quick.refresh_from_db()
        summary = quick.calculation_summary()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(quick.project_name, "Updated Hoodie")
        self.assertEqual(quick.costing_purpose, QuickCosting.PURPOSE_SAMPLE)
        self.assertEqual(summary["total_cost"], Decimal("2000.00"))
        self.assertEqual(summary["revenue"], Decimal("4000.00"))
        self.assertEqual(summary["net_profit_total"], Decimal("1600.00"))

    def test_approval_rejection_and_locked_edit_behavior(self):
        admin = self._admin_user("quick-costing-approval-admin")
        staff = self._costing_user("quick-costing-no-approve", approve=False)
        quick = self._quick_costing(created_by=staff)
        self.client.force_login(admin)

        approve_response = self.client.post(reverse("quick_costing_approve", args=[quick.pk]))
        quick.refresh_from_db()
        self.assertEqual(approve_response.status_code, 302)
        self.assertEqual(quick.status, QuickCosting.STATUS_APPROVED)
        self.assertEqual(quick.approved_by, admin)
        self.assertIsNotNone(quick.approved_at)
        self.assertTrue(quick.is_locked)

        self.client.force_login(staff)
        locked_response = self.client.get(reverse("quick_costing_edit", args=[quick.pk]))
        self.assertEqual(locked_response.status_code, 302)
        self.assertEqual(locked_response["Location"], reverse("quick_costing_detail", args=[quick.pk]))

        self.client.force_login(admin)
        reject_response = self.client.post(reverse("quick_costing_reject", args=[quick.pk]))
        quick.refresh_from_db()
        self.assertEqual(reject_response.status_code, 302)
        self.assertEqual(quick.status, QuickCosting.STATUS_REJECTED)
        self.assertEqual(quick.rejected_by, admin)
        self.assertIsNotNone(quick.rejected_at)
        self.assertIsNone(quick.approved_by)
        self.assertIsNone(quick.approved_at)

    def test_unapproved_quick_costing_cannot_create_quotation(self):
        admin = self._admin_user("quick-costing-unapproved-quote-admin")
        quick = self._quick_costing(created_by=admin)
        self.client.force_login(admin)

        response = self.client.post(reverse("quick_costing_convert_to_quotation", args=[quick.pk]))

        quick.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("quick_costing_detail", args=[quick.pk]))
        self.assertEqual(quick.status, QuickCosting.STATUS_DRAFT)
        self.assertEqual(quick.quotation_number, "")
        self.assertIsNone(quick.quoted_at)

    def test_approved_quick_costing_creates_customer_facing_quotation(self):
        admin = self._admin_user("quick-costing-quote-admin")
        opportunity = self._opportunity()
        quick = self._quick_costing(opportunity=opportunity, created_by=admin)
        self.client.force_login(admin)
        self.client.post(reverse("quick_costing_approve", args=[quick.pk]))

        response = self.client.post(reverse("quick_costing_convert_to_quotation", args=[quick.pk]))
        quick.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("quick_costing_client_quotation", args=[quick.pk]))
        self.assertEqual(quick.status, QuickCosting.STATUS_QUOTED)
        self.assertTrue(quick.quotation_number.startswith("QQT"))
        self.assertEqual(quick.quoted_by, admin)
        self.assertIsNotNone(quick.quoted_at)

        quote_response = self.client.get(reverse("quick_costing_client_quotation", args=[quick.pk]))
        quote_html = quote_response.content.decode("utf-8")
        self.assertEqual(quote_response.status_code, 200)
        self.assertContains(quote_response, "Quotation")
        self.assertContains(quote_response, quick.quotation_number)
        self.assertContains(quote_response, "Selling Price")
        self.assertContains(quote_response, "Currency: Legacy BDT with CAD conversion")
        self.assertContains(quote_response, "Shipping Cost")
        self.assertContains(quote_response, "Total Price")
        self.assertContains(quote_response, "Thank You!")
        self.assertContains(quote_response, "Quotation Status")
        self.assertContains(quote_response, "Approved By")
        self.assertContains(quote_response, admin.username)
        self.assertContains(quote_response, "This quotation was approved through the Quick Costing Approval Workflow.")
        self.assertContains(quote_response, reverse("quick_costing_convert_to_invoice", args=[quick.pk]))
        self.assertContains(quote_response, "Create Invoice")
        self.assertContains(quote_response, "Open Quick Costing")
        self.assertContains(quote_response, reverse("quick_costing_detail", args=[quick.pk]))
        self.assertIn("Quotation", quote_html.split("Workflow Activity Timeline", 1)[1])
        self.assertIn(quick.quotation_number, quote_html)
        self.assertNotContains(quote_response, "Material Cost")
        self.assertNotContains(quote_response, "Production Cost")
        self.assertNotContains(quote_response, "Other Expenses")
        self.assertNotContains(quote_response, "Commission")
        self.assertNotContains(quote_response, "Net Profit")
        self.assertNotContains(quote_response, "Margin Status")

    def test_quick_costing_quotation_rejection_status(self):
        admin = self._admin_user("quick-costing-quote-reject-admin")
        quick = self._quick_costing(created_by=admin)
        self.client.force_login(admin)

        self.client.post(reverse("quick_costing_approve", args=[quick.pk]))
        self.client.post(reverse("quick_costing_convert_to_quotation", args=[quick.pk]))
        self.client.post(reverse("quick_costing_reject", args=[quick.pk]))
        quick.refresh_from_db()

        self.assertEqual(quick.status, QuickCosting.STATUS_REJECTED)
        self.assertEqual(quick.quotation_number, "")
        response = self.client.get(reverse("quick_costing_detail", args=[quick.pk]))
        self.assertContains(response, "Approval Status")
        self.assertContains(response, "Rejected")

    def test_approved_quick_quotation_creates_invoice_and_pdf(self):
        admin = self._admin_user("quick-costing-invoice-admin")
        opportunity = self._opportunity()
        quick = self._quick_costing(opportunity=opportunity, created_by=admin)
        self.client.force_login(admin)

        self.client.post(reverse("quick_costing_approve", args=[quick.pk]))
        self.client.post(reverse("quick_costing_convert_to_quotation", args=[quick.pk]))
        response = self.client.post(reverse("quick_costing_convert_to_invoice", args=[quick.pk]))
        quick.refresh_from_db()

        invoice = Invoice.objects.get(quick_costing=quick)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("invoice_view", args=[invoice.pk]))
        self.assertEqual(quick.status, QuickCosting.STATUS_INVOICED)
        self.assertEqual(invoice.customer, opportunity.customer)
        self.assertEqual(invoice.invoice_market, "north_america")
        self.assertEqual(invoice.invoice_type, "bulk")
        self.assertEqual(invoice.currency, "CAD")
        self.assertEqual(invoice.subtotal, Decimal("16.67"))
        self.assertEqual(invoice.shipping_amount, Decimal("1.11"))
        self.assertEqual(invoice.total_amount, Decimal("17.78"))
        self.assertEqual(invoice.sewing_charge, Decimal("0"))
        self.assertEqual(invoice.other_internal_cost, Decimal("0"))

        invoice_response = self.client.get(reverse("invoice_view", args=[invoice.pk]))
        self.assertContains(invoice_response, invoice.invoice_number)
        self.assertContains(invoice_response, "Quick Costing")
        self.assertContains(invoice_response, opportunity.opportunity_id)
        timeline_html = invoice_response.content.decode("utf-8").split("Workflow Activity Timeline", 1)[1]
        self.assertIn(invoice.invoice_number, timeline_html)

        pdf_response = self.client.get(reverse("invoice_pdf", args=[invoice.pk]))
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response["Content-Type"], "application/pdf")

        quote_response = self.client.get(reverse("quick_costing_client_quotation", args=[quick.pk]))
        self.assertContains(quote_response, "Open Invoice")
        self.assertContains(quote_response, reverse("invoice_view", args=[invoice.pk]))

        duplicate_response = self.client.post(reverse("quick_costing_convert_to_invoice", args=[quick.pk]))
        self.assertEqual(duplicate_response.status_code, 302)
        self.assertEqual(duplicate_response["Location"], reverse("invoice_view", args=[invoice.pk]))
        self.assertEqual(Invoice.objects.filter(quick_costing=quick).count(), 1)

    def test_quick_costing_excel_export(self):
        admin = self._admin_user("quick-costing-excel-admin")
        quick = self._quick_costing(created_by=admin)
        self.client.force_login(admin)

        response = self.client.get(reverse("quick_costing_export_excel", args=[quick.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook.active
        labels = [row[0].value for row in sheet.iter_rows(min_col=1, max_col=1)]
        self.assertIn("Buyer Name", labels)
        self.assertIn("Total Cost", labels)
        self.assertIn("Final Profit After Commission", labels)
        self.assertIn("Status", labels)
        self.assertIn("Approved Date", labels)

    def test_costing_list_shows_quick_workflow_actions_and_status_filter(self):
        admin = self._admin_user("quick-costing-list-actions-admin")
        opportunity = self._opportunity()
        quick = self._quick_costing(
            opportunity=opportunity,
            status=QuickCosting.STATUS_APPROVED,
            costing_purpose=QuickCosting.PURPOSE_SAMPLE,
            created_by=admin,
        )
        bulk = self._quick_costing(
            project_name="Bulk Hoodie",
            opportunity=opportunity,
            status=QuickCosting.STATUS_APPROVED,
            costing_purpose=QuickCosting.PURPOSE_BULK,
            created_by=admin,
        )
        self.client.force_login(admin)

        response = self.client.get(reverse("cost_sheet_list") + "?costing_type=quick&status=approved&purpose=sample")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f"QC-{quick.pk}")
        self.assertNotContains(response, f"QC-{bulk.pk}")
        self.assertContains(response, "Approved")
        self.assertContains(response, "Sample")
        self.assertContains(response, "Meets target")
        self.assertContains(response, reverse("quick_costing_edit", args=[quick.pk]))
        self.assertContains(response, reverse("quick_costing_export_pdf", args=[quick.pk]))
        self.assertContains(response, reverse("quick_costing_export_excel", args=[quick.pk]))
        self.assertContains(response, reverse("opportunity_detail", args=[opportunity.pk]))

        dashboard_response = self.client.get(reverse("cost_sheet_dashboard"))
        reports_response = self.client.get(reverse("cost_sheet_reports"))
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertEqual(reports_response.status_code, 200)
