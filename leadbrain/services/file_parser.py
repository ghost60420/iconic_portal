import csv
import os
import re

from openpyxl import load_workbook

try:
    import xlrd
except Exception:
    xlrd = None


HEADER_ALIASES = {
    "company": "company_name",
    "company_name": "company_name",
    "company name": "company_name",
    "business name": "company_name",
    "name": "company_name",
    "website": "website",
    "site": "website",
    "url": "website",
    "web": "website",
    "website url": "website",
    "company website": "website",
    "email": "email",
    "contact email": "email",
    "contact_email": "email",
    "e mail": "email",
    "phone": "phone",
    "mobile": "phone",
    "telephone": "phone",
    "country": "country",
    "city": "city",
    "linkedin": "linkedin_url",
    "linkedin url": "linkedin_url",
    "notes": "notes",
}


def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value):
    text = _normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def normalize_headers(headers):
    mapping = {}
    for index, header in enumerate(headers or []):
        normalized = _normalize_header(header)
        mapping[index] = HEADER_ALIASES.get(normalized, normalized.replace(" ", "_"))
    return mapping


def _normalize_website(value):
    website = _normalize_text(value)
    if not website:
        return ""
    if "://" not in website and "." in website and " " not in website:
        website = f"https://{website}"
    return website[:200]


def _clean_email(value):
    return _normalize_text(value).lower()


def _score_header_row(values):
    normalized_values = [_normalize_header(value) for value in values if _normalize_text(value)]
    if not normalized_values:
        return -1

    canonical_fields = []
    for value in normalized_values:
        canonical = HEADER_ALIASES.get(value)
        if canonical:
            canonical_fields.append(canonical)

    unique_fields = set(canonical_fields)
    score = len(unique_fields) * 4
    score += len(canonical_fields)
    if "company_name" in unique_fields:
        score += 3
    if "website" in unique_fields:
        score += 3
    if "email" in unique_fields:
        score += 2
    return score


def _detect_header_row(records):
    best_index = None
    best_score = -1

    for index, row in enumerate(records[:25]):
        score = _score_header_row(row)
        if score > best_score:
            best_index = index
            best_score = score

    if best_index is None or best_score < 4:
        raise ValueError(
            "No header row could be detected. Include a header with columns like Company, Website, or Email."
        )

    header_row = records[best_index]
    return best_index, header_row, normalize_headers(header_row)


def _preview_row(row):
    return {
        "row_number": row.get("row_number", 0),
        "company_name": row.get("company_name", ""),
        "website": row.get("website", ""),
        "email": row.get("email", ""),
        "phone": row.get("phone", ""),
        "country": row.get("country", ""),
        "city": row.get("city", ""),
    }


def extract_company_row(row, row_number):
    cleaned = {
        "row_number": row_number,
        "company_name": _normalize_text(row.get("company_name", "")),
        "website": _normalize_website(row.get("website", "")),
        "email": _clean_email(row.get("email", "")),
        "phone": _normalize_text(row.get("phone", "")),
        "country": _normalize_text(row.get("country", "")),
        "city": _normalize_text(row.get("city", "")),
        "raw_row_json": {str(key): value for key, value in (row or {}).items()},
    }
    return cleaned


def _build_rows_from_records(records):
    if not records:
        raise ValueError("The uploaded file is empty.")

    header_index, header_row, header_map = _detect_header_row(records)
    parsed_rows = []
    blank_rows = 0
    detected_columns = []

    for index, value in enumerate(header_row):
        header_value = _normalize_text(value)
        if not header_value:
            continue
        detected_columns.append(
            {
                "index": index,
                "source": header_value,
                "canonical": header_map.get(index, f"column_{index + 1}"),
            }
        )

    for raw_index, values in enumerate(records[header_index + 1 :], start=header_index + 2):
        row_dict = {}
        raw_row = {}
        for index, value in enumerate(values):
            header_value = header_row[index] if index < len(header_row) else f"column_{index + 1}"
            raw_row[str(header_value)] = value
            canonical_key = header_map.get(index, f"column_{index + 1}")
            row_dict[canonical_key] = _normalize_text(value)

        if not any(_normalize_text(value) for value in row_dict.values()):
            blank_rows += 1
            continue

        row_dict["raw_row_json"] = raw_row
        parsed_rows.append(extract_company_row(row_dict, raw_index))

    if not parsed_rows:
        raise ValueError("No usable rows were found in the uploaded file.")
    return {
        "rows": parsed_rows,
        "source_row_count": len(parsed_rows),
        "blank_rows": blank_rows,
        "header_row_number": header_index + 1,
        "detected_columns": detected_columns,
        "sample_rows": [_preview_row(row) for row in parsed_rows[:5]],
    }


def _parse_csv(file_path):
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(file_path, "r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                rows = [row for row in reader]
            return _build_rows_from_records(rows)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError("The CSV file could not be read.") from last_error


def _parse_xlsx(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([_normalize_text(value) for value in row])
    return _build_rows_from_records(rows)


def _parse_xls(file_path):
    if xlrd is None:
        raise ValueError("XLS import is not available on this server. Please upload CSV or XLSX.")

    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)
    rows = []
    for row_index in range(sheet.nrows):
        rows.append([_normalize_text(sheet.cell_value(row_index, col)) for col in range(sheet.ncols)])
    return _build_rows_from_records(rows)


def parse_uploaded_file(file_path):
    return parse_uploaded_file_report(file_path)["rows"]


def parse_uploaded_file_report(file_path):
    ext = os.path.splitext(file_path or "")[1].lower()
    if ext == ".csv":
        return _parse_csv(file_path)
    if ext == ".xlsx":
        return _parse_xlsx(file_path)
    if ext == ".xls":
        return _parse_xls(file_path)
    raise ValueError("Unsupported file type. Please upload CSV, XLSX, or XLS.")
